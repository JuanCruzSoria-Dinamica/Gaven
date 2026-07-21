"""Genera data/proveedor_objetivo_lookup.csv a partir del Excel comercial.

La columna 'marca_linea' es el "PROVEEDOR OBJETIVO" del tablero (hoja
'Formulador', columna GB). Acá se REPLICA la fórmula del Excel en Python y se
precalcula una sola vez por CÓDIGO de artículo. Como grupo, familia, línea y
proveedor son atributos fijos del artículo, el resultado es constante por
código, así que el pipeline después joina por idArticulo (== "Código de
Artículo" del ERP).

Uso:
    python build_lookup_proveedor_objetivo.py "2026 Conversor Comercial.xlsx"

Requisitos: openpyxl. Regenerar cuando el Excel sume artículos nuevos.
La replicación fue validada 1:1 contra los valores ya calculados por el Excel
(0 diferencias en las 24.976 filas reales de t_Base).
"""

import csv
import os
import sys

import openpyxl
from openpyxl.utils import column_index_from_string as ci

# Columnas de la hoja 'Formulador' (tabla t_Base)
COL_COD = "CB"      # Codigo de Articulo
COL_DESC = "CC"     # Descripcion de Articulo
COL_GRUPO = "DA"    # Descripción ARTICULO GRUPO
COL_LINEA = "DE"    # Descripción ARTICULO LINEA
COL_PROV = "DG"     # Descripción PROVEEDORES
COL_FAM = "DM"      # Descripción ARTICULO FAMILIA


def _cargar_tfood(wb):
    """T_FOOD (hoja 'TABLAS ', rango AN6:AR116): código -> CLIENTE MCCAIN (col AR).
    Se usa sólo para los artículos del proveedor FRIAR S. A."""
    ws = wb["TABLAS "]
    tfood = {}
    for r in ws.iter_rows(min_row=7, max_row=116,
                          min_col=ci("AN"), max_col=ci("AR"), values_only=True):
        if r[0] is not None:
            tfood[int(r[0])] = r[4]
    return tfood


def proveedor_objetivo(prov, grupo, fam, linea, cod, tfood):
    """Réplica exacta de la fórmula Formulador!GB (PROVEEDOR OBJETIVO)."""
    prov = (prov or "").strip()
    grupo = grupo or ""
    fam = fam or ""
    linea = linea or ""
    if prov == "GRANJA TRES ARROYOS SOCIEDAD ANONIMA":
        return f"GRANJA TRES ARROYOS SOCIEDAD ANONIMA {grupo}".strip()
    if prov == "GARCIA HNOS AGROINDUSTRIAL SRL":
        return f"GARCIA HNOS AGROINDUSTRIAL SRL {fam}".strip()
    if prov == "MC CAIN ARGENTINA SA" and linea in ("SIBARITA RETAIL", "MC CAIN RETAIL"):
        return "MC CAIN RETAIL"
    if prov == "MC CAIN ARGENTINA SA" and linea in ("MC CAIN FOOD SERVICE", "MC CAIN VEGETALES"):
        return "MC CAIN FOOD"
    if prov == "ERNESTO RODRIGUEZ E HIJOS SA" and fam == "QUESOS":
        return f"{prov} LACTEOS"
    if prov == "ERNESTO RODRIGUEZ E HIJOS SA":
        return f"{prov} {fam}".strip()
    if prov == "ELCOR S.A" and cod in (11515, 20101):
        return "ELCOR S.A FOOD"
    if prov == "ELCOR S.A":
        return "ELCOR S.A RETAIL"
    if prov == "FRIAR S. A":
        v = tfood.get(int(cod)) if cod is not None else None
        return v if v is not None else "NO"
    return prov


def main(xlsx_path, out_path=None):
    if out_path is None:
        out_path = os.path.join(os.path.dirname(__file__), "data",
                                "proveedor_objetivo_lookup.csv")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    tfood = _cargar_tfood(wb)
    ws = wb["Formulador"]

    rec = {}  # código -> fila (constante por código; se toma la 1ra aparición)
    for row in ws.iter_rows(min_row=3, max_col=ci(COL_PROV) if ci(COL_PROV) > ci(COL_FAM) else ci(COL_FAM), values_only=True):
        cod = row[ci(COL_COD) - 1]
        if cod is None or int(cod) == 0:
            continue
        cod = int(cod)
        if cod in rec:
            continue
        prov = row[ci(COL_PROV) - 1]
        grupo = row[ci(COL_GRUPO) - 1]
        fam = row[ci(COL_FAM) - 1]
        linea = row[ci(COL_LINEA) - 1]
        rec[cod] = {
            "idArticulo": cod,
            "dsArticulo": row[ci(COL_DESC) - 1],
            "marca_linea": proveedor_objetivo(prov, grupo, fam, linea, cod, tfood),
            "proveedor_ref": prov,
            "grupo": grupo,
            "familia": fam,
            "linea": linea,
        }

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["idArticulo", "dsArticulo", "marca_linea",
                                          "proveedor_ref", "grupo", "familia", "linea"])
        w.writeheader()
        for cod in sorted(rec):
            w.writerow(rec[cod])
    print(f"OK: {len(rec)} códigos -> {out_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
