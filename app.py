"""
app.py
------
Panel de ventas (Gaven). SOLO presentación (todos los meses de 2026).

NO llama al API ni hace el procesamiento pesado: lee el archivo que dejó
data_pipeline.py (data/ventas_actualizadas.parquet) y muestra todo.

Correr local:   streamlit run app.py
"""

import io
import os
import html
import json
import time
import calendar
import datetime as dt

import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px

import data_pipeline as dp


# ---------------------------------------------------------------------------
# Configuración de la página
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="Panel de Ventas · Gaven",
    layout="wide",
)

# --- Estilos (paleta del tablero de referencia) ----------------------------
st.markdown(
    """
    <style>
      :root{
        --verde:#00b87a; --azul:#2a8ed4; --naranja:#f59e0b;
        --violeta:#a78bfa; --rojo:#f87171;
        --sf:#111827; --sf2:#1a2332; --border:#2a3a50; --tx2:#94a3b8;
      }
      .block-container{padding-top:2.2rem; max-width:1500px;}
      h1{font-weight:700; letter-spacing:-.5px;}
      /* Tarjetas de métricas */
      [data-testid="stMetric"]{
        background:var(--sf); border:1px solid var(--border);
        border-radius:12px; padding:14px 16px;
      }
      [data-testid="stMetricLabel"]{color:var(--tx2); font-size:.78rem;}
      [data-testid="stMetricValue"]{font-weight:700;}
      /* Tabs */
      .stTabs [data-baseweb="tab-list"]{gap:4px; border-bottom:1px solid var(--border);}
      .stTabs [data-baseweb="tab"]{
        border-radius:8px 8px 0 0; padding:8px 16px; font-weight:500;
      }
      .stTabs [aria-selected="true"]{
        background:var(--verde); color:#04221a !important;
      }
      /* FIX: mantener ocultas las solapas inactivas.
         Streamlit manda el contenido de TODAS las solapas al navegador y solo
         esconde las inactivas con CSS. Cuando un widget dentro de una solapa
         (ej. el evolutivo) dispara un rerun, esa regla de ocultamiento a veces
         se pierde y todo el contenido aparece apilado en todas las solapas.
         Forzamos que los paneles inactivos ([hidden]) sigan ocultos. */
      .stTabs [data-baseweb="tab-panel"][hidden],
      .stTabs [role="tabpanel"][hidden]{
        display:none !important;
      }
      /* Sidebar vacía: la ocultamos (los filtros van arriba) */
      section[data-testid="stSidebar"]{display:none;}
      /* Barra de filtros (contenedor con borde) */
      [data-testid="stVerticalBlockBorderWrapper"]{
        background:var(--sf); border-radius:12px;
      }
      /* Subtítulos */
      h3{color:#cbd5e1; font-weight:600; letter-spacing:-.2px;}
      /* Los objetivos se tipean, no se suben de a pasos: sacamos los botones
         − / + del number_input (subir 95.000 kg de a 1.000 no tiene sentido).
         Se cubren los test-id nuevos y las clases viejas por si cambia la
         versión de Streamlit. El input sigue aceptando las flechas del
         teclado. */
      [data-testid="stNumberInputStepUp"],
      [data-testid="stNumberInputStepDown"],
      [data-testid="stNumberInput"] button{
        display:none !important;
      }
      /* Tarjetas de las lecturas de mesa chica (solapa Alertas) */
      .ins-card{
        background:var(--sf); border:1px solid var(--border);
        border-left:4px solid var(--azul);
        border-radius:12px; padding:12px 14px; height:100%;
        margin-bottom:12px;
      }
      .ins-card.ok{border-left-color:var(--verde);}
      .ins-card.riesgo{border-left-color:var(--rojo);}
      .ins-card .ins-tit{
        color:var(--tx2); font-size:.72rem; text-transform:uppercase;
        letter-spacing:.5px; font-weight:600; margin-bottom:2px;
      }
      .ins-card .ins-prot{
        font-size:.98rem; font-weight:700; line-height:1.25; margin-bottom:2px;
      }
      .ins-card .ins-val{font-size:1.15rem; font-weight:700; color:var(--verde);}
      .ins-card.riesgo .ins-val{color:var(--rojo);}
      .ins-card .ins-det{
        color:#cbd5e1; font-size:.8rem; line-height:1.4; margin-top:8px;
      }
      /* Un renglón por dato, con una línea divisoria tenue entre ellos para
         que se lean de un vistazo sin buscar los separadores. */
      .ins-card .ins-det > div{
        padding:4px 0; border-top:1px solid var(--border);
      }
      .ins-card .ins-det > div:first-child{border-top:none; padding-top:0;}
      .ins-card .ins-det > div:last-child{padding-bottom:0;}
    </style>
    """,
    unsafe_allow_html=True,
)

PARQUET_PATH = dp.PARQUET_PATH
META_PATH = dp.META_PATH


# ---------------------------------------------------------------------------
# Helpers de formato
# ---------------------------------------------------------------------------

def fmt_money(x):
    try:
        return f"$ {x:,.0f}".replace(",", ".")
    except (TypeError, ValueError):
        return x


def fmt_kg(x):
    try:
        return f"{x:,.0f} kg".replace(",", ".")
    except (TypeError, ValueError):
        return x


# ---------------------------------------------------------------------------
# Lectura de datos locales (se cachea la LECTURA del archivo, no el API).
# La clave de caché incluye el mtime: si el pipeline reescribe el parquet,
# el mtime cambia y la caché se invalida sola.
# ---------------------------------------------------------------------------

def _mtime_acuerdos():
    """mtime del almacén de acuerdos McCain (0 si aún no existe). Entra en la
    clave de caché: al subir un Excel nuevo, las cachés se invalidan solas."""
    try:
        return os.path.getmtime(dp.ACUERDOS_PATH)
    except OSError:
        return 0


def _mtime_metas():
    """mtime del parquet de metas (0 si aún no existe). Mismo criterio que
    _mtime_acuerdos(): al guardar objetivos nuevos, todo lo que se cachea a
    partir de las metas se invalida solo."""
    try:
        return os.path.getmtime(dp.METAS_PATH)
    except OSError:
        return 0


# max_entries: tope de versiones vivas en caché. Sin esto, cada vez que el
# pipeline reescribe el parquet cambia el mtime -> nueva entrada, y la vieja
# NUNCA se libera. Eso es lo que hacía que la app se cayera sola después de un
# rato en el servidor (se queda sin RAM y Streamlit Cloud mata el proceso).
@st.cache_data(show_spinner="Leyendo datos...", max_entries=2)
def cargar_datos_local(mtime, mtime_acuerdos=0):
    df = pd.read_parquet(PARQUET_PATH)
    # 'marca_linea' es una columna DERIVADA del lookup por código
    # (data/proveedor_objetivo_lookup.csv). Se recalcula siempre al leer para
    # que la clasificación refleje el lookup vigente aunque el parquet guardado
    # traiga valores viejos. Es barato (map por idArticulo).
    df = dp.agregar_marca_linea(df)
    # Ajuste de costo por acuerdos McCain (descuentos que Chess no informa).
    # El parquet guarda el costo CRUDO; acá se aplica el descuento vigente,
    # así subir un Excel nuevo corrige CM y CM% al instante, sin re-correr
    # el pipeline. Agrega la columna 'ajuste_mccain' (auditoría).
    df = dp.aplicar_acuerdos(df)
    return df


@st.cache_data(show_spinner="Leyendo serie histórica...", max_entries=2)
def cargar_serie(mtime, mtime_parquet=0, mtime_acuerdos=0):
    """Lee la serie mensual agregada (data/serie_mensual.parquet).
    La clave de caché es el mtime: si el pipeline reescribe la serie, se
    invalida sola (mismo patrón que cargar_datos_local).

    Los meses que también están en el parquet de DETALLE se recalculan desde
    el detalle con el costo ya ajustado por acuerdos McCain, para que la
    evolución de CM/CM% coincida con el resto del tablero. Los meses viejos
    (solo-serie, ej. 2025) quedan tal cual los dejó el backfill."""
    serie = pd.read_parquet(dp.SERIE_PATH)
    if dp.cargar_acuerdos().empty:
        return serie
    det = cargar_datos_local(mtime_parquet, mtime_acuerdos)
    nuevos = dp.agregar_serie(det)
    if nuevos.empty:
        return serie
    serie = serie[~serie["anio_mes"].isin(set(nuevos["anio_mes"]))]
    return (
        pd.concat([serie, nuevos], ignore_index=True)
        .sort_values(["anio_mes"] + dp.SERIE_GRANO[1:])
        .reset_index(drop=True)
    )


# max_entries chico a propósito: la clave combina nivel × canales elegidos ×
# mtimes × fecha. Sin tope, cada combinación que toca el usuario deja un
# DataFrame en memoria para siempre.
@st.cache_data(show_spinner="Armando el evolutivo...", max_entries=8)
def cargar_evolutivo(nivel, canales, mtime, mtime_acuerdos=0, mtime_metas=0,
                     hoy=None):
    """Evolutivo de objetivo vs. real mes a mes (ver dp.evolutivo_metas()).

    Se cachea porque recorre TODOS los meses del detalle y proyecta el mes
    abierto vendedor por vendedor: es la cuenta más cara de la solapa y no
    cambia hasta que entran ventas nuevas o se guarda un objetivo. Los mtime
    entran en la clave para que se invalide sola en los dos casos.

    `canales` llega como tupla porque la clave de caché tiene que ser
    hasheable.
    """
    df = cargar_datos_local(mtime, mtime_acuerdos)
    return dp.evolutivo_metas(
        df, dp.cargar_metas(),
        nivel=nivel,
        canales=list(canales) if canales else None,
        anio=dp.ANIO,
        hoy=hoy,
    )


@st.cache_data(show_spinner="Leyendo IPC (INDEC)...")
def cargar_ipc(_mtime=None):
    """Devuelve el IPC del INDEC. Usa el archivo que deja el pipeline; si todavía
    no existe (ej. antes de la primera corrida del cron), intenta bajarlo una vez.
    La clave de caché es el mtime del archivo: cuando el pipeline reescribe el
    IPC, la caché se invalida sola (igual que la serie). Así nunca queda
    'pegado' un IPC vacío."""
    ipc = dp.cargar_ipc()
    if ipc.empty:
        try:
            ipc = dp.descargar_ipc()
        except Exception:
            pass
    return ipc


def leer_metadata():
    try:
        with open(META_PATH, encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError):
        return {}


# ---------------------------------------------------------------------------
# Rango de fechas: cualquier mes de 2026 con datos en el parquet.
# El pipeline mantiene el detalle de TODO el año por upsert mensual, así que
# acá solo listamos los meses disponibles y armamos el rango del elegido.
# ---------------------------------------------------------------------------

MESES_ES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
            "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def rango_mes(anio_mes, hoy=None):
    """(desde, hasta) del mes calendario 'YYYY-MM'. El mes en curso se corta
    en hoy; los meses cerrados van del día 1 al último día del mes."""
    hoy = hoy or dt.date.today()
    anio, mes = map(int, str(anio_mes).split("-"))
    desde = dt.date(anio, mes, 1)
    ultimo = dt.date(anio, mes, calendar.monthrange(anio, mes)[1])
    return desde, min(ultimo, hoy)


def etiqueta_mes(anio_mes, hoy=None):
    """'2026-07' -> 'Julio 2026 (Actual)' / '2026-03' -> 'Marzo 2026'."""
    hoy = hoy or dt.date.today()
    anio, mes = map(int, str(anio_mes).split("-"))
    lbl = f"{MESES_ES[mes - 1]} {anio}"
    if anio_mes == hoy.strftime("%Y-%m"):
        lbl += " (Actual)"
    return lbl


def meses_disponibles(df, anio=None):
    """Meses 'YYYY-MM' del año con datos en el parquet, más reciente primero."""
    anio = anio or dp.ANIO
    f = pd.to_datetime(df["fechaComprobate"], errors="coerce").dropna()
    f = f[f.dt.year == anio]
    return sorted(f.dt.strftime("%Y-%m").unique(), reverse=True)


# ---------------------------------------------------------------------------
# Login + roles
# ---------------------------------------------------------------------------
# Dos roles, sin base de datos de usuarios. Las credenciales viven en
# .streamlit/secrets.toml (sección [acceso]), NO en este archivo.
#   - "dueno"      -> ve todo, incluida Contribución marginal y CM %.
#   - "supervisor" -> ve todo MENOS Contribución marginal y CM %.
# El rol se guarda en st.session_state y sobrevive a las re-ejecuciones.

def _login():
    """Muestra el login y corta la ejecución hasta que el rol esté seteado."""
    if "rol" not in st.session_state:
        st.session_state.rol = None
    if st.session_state.rol is not None:
        # Pantalla de carga de 1 segundo, solo justo después de loguearse.
        if st.session_state.pop("_cargando_login", False):
            st.markdown(
                """
                <style>
                  @keyframes girar{to{transform:rotate(360deg);}}
                  /* Overlay a pantalla completa: tapa el contenido "viejo"
                     que Streamlit deja visible mientras corre el script. */
                  .carga-login{
                    position:fixed; inset:0; z-index:999999;
                    background:#0a0e17;
                    display:flex; flex-direction:column; align-items:center;
                    justify-content:center; gap:18px;
                  }
                  .carga-login .aro{
                    width:44px; height:44px; border-radius:50%;
                    border:4px solid var(--border);
                    border-top-color:var(--verde);
                    animation:girar .8s linear infinite;
                  }
                  .carga-login p{color:var(--tx2); font-size:.9rem; margin:0;}
                </style>
                <div class="carga-login">
                  <div class="aro"></div>
                  <p>Cargando panel…</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
            time.sleep(1)
            st.rerun()
        return  # ya está logueado, seguimos con el tablero

    try:
        cred = st.secrets["acceso"]
    except Exception:
        st.error(
            "Falta la sección [acceso] en .streamlit/secrets.toml. "
            "Agregá usuarios y contraseñas para poder entrar."
        )
        st.stop()

    # Login centrado en una "cajita" angosta (como cualquier sitio web):
    # tres columnas y el formulario va en la del medio.
    st.markdown(
        """
        <style>
          /* Quitamos el borde propio del form: la "caja" la pone el
             contenedor con borde de afuera, así no se duplica. */
          [data-testid="stForm"]{border:0; padding:0;}
          .login-head{text-align:center; margin:0 0 16px;}
          .login-head h2{margin:0; font-weight:700; letter-spacing:-.3px;
            font-size:1.25rem;}
          .login-head p{color:var(--tx2); font-size:.85rem; margin:.25rem 0 0;}
        </style>
        """,
        unsafe_allow_html=True,
    )
    _l, _c, _r = st.columns([1.4, 1, 1.4])
    with _c:
        st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(
                "<div class='login-head'>"
                "<h2>Panel de Ventas · Gaven</h2>"
                "<p>Iniciá sesión para continuar</p>"
                "</div>",
                unsafe_allow_html=True,
            )
            with st.form("login"):
                usuario = st.text_input("Usuario")
                pwd = st.text_input("Contraseña", type="password")
                entrar = st.form_submit_button(
                    "Entrar", type="primary", use_container_width=True
                )
        if entrar:
            if (usuario == cred.get("usuario_duenos")
                    and pwd == cred.get("password_duenos")):
                st.session_state.rol = "dueno"
                st.session_state._cargando_login = True
                st.rerun()
            elif (usuario == cred.get("usuario_supervisores")
                    and pwd == cred.get("password_supervisores")):
                st.session_state.rol = "supervisor"
                st.session_state._cargando_login = True
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos.")
    st.stop()  # mientras no haya rol válido, no se renderiza nada del tablero


_login()

# A partir de acá hay un rol válido en sesión.
# mostrar_cm es la llave maestra: si es False, los números de Contribución
# marginal y CM % no se calculan ni se muestran en NINGÚN lado del tablero.
mostrar_cm = st.session_state.rol == "dueno"


# ---------------------------------------------------------------------------
# Carga + guardas
# ---------------------------------------------------------------------------

# Encabezado con el rol activo y botón para cerrar sesión.
_ct, _cu = st.columns([4, 1])
_ct.title("Panel de Ventas · Gaven")
_rol_label = "Dueño" if st.session_state.rol == "dueno" else "Supervisor"
_cu.markdown(
    f"<div style='text-align:right;color:var(--tx2);font-size:.8rem;"
    f"padding-top:1rem'>Sesión: <b>{_rol_label}</b></div>",
    unsafe_allow_html=True,
)
if _cu.button("Cerrar sesión", use_container_width=True):
    st.session_state.rol = None
    st.rerun()

if not os.path.exists(PARQUET_PATH):
    st.warning(
        "Todavía no hay datos cargados.\n\n"
        "Ejecutá primero el pipeline para generar el archivo:\n\n"
        "```\npython data_pipeline.py\n```"
    )
    st.stop()

df = cargar_datos_local(os.path.getmtime(PARQUET_PATH), _mtime_acuerdos())

# Copia del detalle del AÑO COMPLETO antes de recortarlo por período. Es la
# base del universo con el que se calcula la cobertura (ver más abajo):
# `df` se pisa enseguida con el mes elegido y después ya no se puede recuperar.
df_anio = df


# ---------------------------------------------------------------------------
# Barra de filtros (arriba). Todos los filtros son selectores y aplican a
# TODAS las solapas (filtro global), igual que el tablero de referencia.
# ---------------------------------------------------------------------------

def opciones(serie):
    """Lista ordenada de valores únicos no vacíos para un multiselect."""
    vals = (
        serie.dropna().astype(str).str.strip()
        .replace({"": None, "0": None}).dropna().unique().tolist()
    )
    return sorted(vals)


# Cada filtro es una tupla: (etiqueta, columna). Solo se muestran los que
# realmente tienen datos en el período.
FILTROS = [
    ("Canal", "dsCanalMkt"),
    ("Subcanal", "dsSubcanalMKT"),
    ("Región", "region"),
    ("Vendedor", "dsVendedor"),
    ("Marca / Línea", "marca_linea"),
    ("Cliente", "nombreCliente"),
]

with st.container(border=True):
    # Fila 1: período + última actualización
    f1a, f1c = st.columns([2, 2])
    _meses_disp = meses_disponibles(df)
    if not _meses_disp:
        st.warning(
            f"El parquet no tiene datos de {dp.ANIO}. "
            "Corré el pipeline: `python data_pipeline.py`"
        )
        st.stop()
    mes_sel = f1a.selectbox(
        "Período", _meses_disp, index=0, format_func=etiqueta_mes,
        help="Todos los meses de 2026 con datos. El pipeline trae los meses "
             "faltantes una sola vez y después solo actualiza el mes en curso "
             "y el anterior.",
    )
    es_mes_actual = mes_sel == dt.date.today().strftime("%Y-%m")
    desde, hasta = rango_mes(mes_sel)

    # El mes en curso se corta en la última fecha CON DATOS, no en hoy: si el
    # pipeline corrió con --hasta (corte a un día específico), la etiqueta
    # "Período" refleja ese corte real y no la fecha de hoy.
    if es_mes_actual:
        _ult_dato = df["fechaComprobate"].max()
        if pd.notna(_ult_dato):
            hasta = min(hasta, _ult_dato.date())

    meta = leer_metadata()
    ultima = meta.get("ultima_actualizacion", "—")
    f1c.markdown(
        f"<div style='text-align:right;color:var(--tx2);font-size:.8rem;"
        f"padding-top:1.9rem'>Última actualización: {ultima}</div>",
        unsafe_allow_html=True,
    )

    # df del período (base para construir las opciones de los selectores)
    fecha = df["fechaComprobate"]
    df_periodo = df[
        (fecha >= pd.Timestamp(desde))
        & (fecha < pd.Timestamp(hasta) + pd.Timedelta(days=1))
        & (fecha.dt.year == dp.ANIO)
    ].copy()

    # Fila 2: un selector por dimensión (solo las que tienen datos).
    # Filtros EN CASCADA: las opciones de cada selector se calculan sobre el
    # df ya filtrado por los OTROS selectores. Así, si filtrás por "food
    # service", el selector de Vendedor solo ofrece los que vendieron eso.
    seleccion = {}
    if not df_periodo.empty:
        disponibles = [
            (et, col) for et, col in FILTROS
            if col in df_periodo.columns and opciones(df_periodo[col])
        ]
        if disponibles:
            # Selecciones de la corrida anterior (Streamlit re-ejecuta en cada
            # interacción): sirven de base para armar las opciones cruzadas.
            sel_prev = {
                col: st.session_state.get(f"filtro_{col}", [])
                for _, col in disponibles
            }

            def _df_filtrado_excepto(col_excluida):
                """df del período filtrado por todos los selectores menos uno."""
                d = df_periodo
                for c, vals in sel_prev.items():
                    if c == col_excluida or not vals:
                        continue
                    d = d[d[c].astype(str).str.strip().isin(vals)]
                return d

            cols = st.columns(len(disponibles))
            for i, (etiqueta, col) in enumerate(disponibles):
                opts = opciones(_df_filtrado_excepto(col)[col])
                key = f"filtro_{col}"
                # Si algún valor elegido ya no es válido (porque otro filtro lo
                # excluyó), lo sacamos del estado para evitar el error de
                # Streamlit "default value not in options".
                if key in st.session_state:
                    st.session_state[key] = [
                        v for v in st.session_state[key] if v in opts
                    ]
                seleccion[col] = cols[i].multiselect(
                    etiqueta, opts, key=key, placeholder="Todos",
                )

n_filtros = sum(1 for v in seleccion.values() if v)
chip = f"  ·  {n_filtros} filtro(s) activo(s)" if n_filtros else "  ·  sin filtros"
st.caption(f"Período: {desde:%d/%m/%Y} → {hasta:%d/%m/%Y}{chip}")
st.divider()

if df_periodo.empty:
    st.warning("No hay datos para el período seleccionado.")
    st.stop()

# Aplica los filtros seleccionados (los vacíos no filtran nada)
df = df_periodo
for col, valores in seleccion.items():
    if valores:
        df = df[df[col].astype(str).str.strip().isin(valores)]

if df.empty:
    st.warning("No hay datos que cumplan con los filtros seleccionados.")
    st.stop()

# --- Universo del año (denominador de la cobertura) ------------------------
# Mismos filtros de dimensión que `df`, pero SIN recortar por período: todo
# 2026. Contra este universo se mide qué porción de la cartera y del surtido
# se tocó en el mes elegido (dp.agregar_cobertura / dp.cobertura_total).
df_universo = df_anio[df_anio["fechaComprobate"].dt.year == dp.ANIO]
for col, valores in seleccion.items():
    if valores:
        df_universo = df_universo[
            df_universo[col].astype(str).str.strip().isin(valores)
        ]

# ---------------------------------------------------------------------------
# Formatos de tablas reutilizables
# ---------------------------------------------------------------------------

def fmt_pct(x):
    try:
        return f"{x:.1f} %"
    except (TypeError, ValueError):
        return x


# Columnas "estándar" que devuelve dp.agrupar_dim, con sus nombres lindos
COLS_DIM = {
    "kilos": "Kilos", "subtotalNeto": "Facturación", "cm": "Contribución",
    "cm_pct": "CM %", "precio_kg": "$/kg", "clientes": "Clientes",
    "skus": "SKUs", "skus_por_cliente": "SKUs/Cliente",
    "share_fc": "Share FC %", "share_kg": "Share Kg %",
    "share_cm": "Share CM %",
    # Cobertura: % de la cartera / del surtido del año que se tocó en el
    # período. Las columnas solo existen si se pasó por dp.agregar_cobertura.
    "universo_clientes": "Cartera", "universo_skus": "Surtido",
    "cob_clientes": "Cob. clientes %", "cob_skus": "Cob. SKUs %",
}
FMT_DIM = {
    "Kilos": fmt_kg, "Facturación": fmt_money, "Contribución": fmt_money,
    "CM %": fmt_pct, "$/kg": fmt_money, "Share FC %": fmt_pct, "Share Kg %": fmt_pct,
    "Share CM %": fmt_pct,
    "SKUs/Cliente": lambda x: f"{x:,.1f}".replace(",", "."),
    "Cob. clientes %": fmt_pct, "Cob. SKUs %": fmt_pct,
}


def tabla_dim(g, dim_label, dim_col, mostrar_skus=False,
              mostrar_skus_cliente=False):
    """Renderiza un resumen de dp.agrupar_dim como tabla formateada.

    mostrar_skus=True agrega la columna 'SKUs' (cantidad de productos únicos
    que maneja cada fila de la dimensión).
    mostrar_skus_cliente=True agrega 'SKUs/Cliente' (productos únicos
    promedio por cliente).

    Las columnas de cobertura (cartera / surtido del año y sus %) se muestran
    solas si `g` viene de dp.agregar_cobertura; si no, no aparecen. Van
    pegadas a su numerador: Clientes → Cartera → Cob. clientes %."""
    cols = [dim_col, "kilos", "subtotalNeto", "share_fc", "cm", "share_cm",
            "cm_pct", "precio_kg", "clientes", "universo_clientes",
            "cob_clientes"]
    if mostrar_skus:
        cols += ["skus", "universo_skus", "cob_skus"]
    if mostrar_skus_cliente:
        cols.append("skus_por_cliente")
    # Supervisores no ven Contribución ni CM %: se quitan las columnas.
    if not mostrar_cm:
        cols = [c for c in cols if c not in ("cm", "share_cm", "cm_pct")]
    cols = [c for c in cols if c in g.columns]
    t = g[cols].rename(columns={dim_col: dim_label, **COLS_DIM})
    st.dataframe(
        t.style.format(FMT_DIM), use_container_width=True, hide_index=True,
    )
    # Se devuelve la tabla ya renombrada (y sin CM si el rol no la ve) para
    # poder incluirla en el Excel descargable de la solapa.
    return t


# ---------------------------------------------------------------------------
# Descarga en Excel (un botón por solapa)
# ---------------------------------------------------------------------------
# El botón nativo de las tablas de Streamlit solo baja CSV crudo. Estos
# helpers arman un .xlsx real (openpyxl) con una hoja por tabla, números como
# números y anchos de columna razonables. Las tablas que reciben ya vienen
# renombradas y filtradas por rol (sin CM para supervisores).

# Formatos numéricos de Excel (el separador de miles/decimales lo resuelve
# Excel según la configuración regional del usuario: con locale es-AR queda
# "$ 1.234.567", "12,5 %" y "1.234.567").
XL_MONEY = '"$" #,##0'          # plata: signo $ y sin decimales
XL_PCT = '#,##0.0"%"'           # porcentaje: 1 decimal (valores ya en 0-100)
XL_KG = '#,##0'                 # kilos: punto de miles, sin decimales
XL_INT = '#,##0'
XL_DEC1 = '#,##0.0'

# Columnas de plata que no se detectan por el nombre (no llevan "$").
_COLS_MONEY = {
    "Facturación", "Facturación neta", "Contribución",
    "Contribución marginal", "Ajuste aplicado", "Monetario",
}


def _formato_col(nombre, serie):
    """Devuelve el number_format de Excel según el nombre de la columna.

    Reglas: todo lo que sea plata va con "$" y sin decimales; todo lo que sea
    porcentaje con "%" y 1 decimal; los kilos con punto de miles."""
    n = str(nombre).strip()
    bajo = n.lower()
    if "%" in n:
        return XL_PCT
    if (n in _COLS_MONEY or "$" in n or bajo.startswith("facturación")
            or bajo.startswith("contribución") or bajo.startswith("ajuste")
            or bajo.startswith("precio") or bajo.startswith("ticket")):
        return XL_MONEY
    if bajo.startswith("kilos") or "(kg)" in bajo or bajo.endswith(" kg"):
        return XL_KG
    if pd.api.types.is_integer_dtype(serie):
        return XL_INT
    # Conteos que quedaron como float (clientes, SKUs, compras...): sin
    # decimales, pero igual con separador de miles.
    try:
        s = pd.to_numeric(serie, errors="coerce").dropna()
        if len(s) and (s % 1 == 0).all():
            return XL_INT
    except (TypeError, ValueError):
        pass
    return XL_DEC1


def _excel_bytes(hojas, formatos=None):
    """hojas: dict {nombre_hoja: DataFrame} -> bytes de un .xlsx.

    formatos: dict opcional {nombre_hoja: spec} para forzar el formato cuando
    el nombre de la columna no alcanza para deducirlo. `spec` puede ser un
    number_format (se aplica a todas las columnas numéricas de la hoja, ej. la
    pivote por mes) o un dict {columna: number_format}; el valor por columna
    puede además ser una lista con un formato por fila (ej. la hoja KPIs, que
    mezcla plata, kilos y % en la misma columna)."""
    from openpyxl.utils import get_column_letter

    formatos = formatos or {}
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        for nombre, t in hojas.items():
            if t is None or len(t) == 0:
                continue
            nombre = str(nombre)[:31]  # límite de Excel
            spec = formatos.get(nombre)
            t.to_excel(w, sheet_name=nombre, index=False)
            ws = w.sheets[nombre]
            for i, c in enumerate(t.columns, 1):
                letra = get_column_letter(i)
                # Fechas: formato corto dd/mm/aaaa (sin la hora 00:00:00,
                # que hace que Excel muestre '####' si la columna es angosta).
                if pd.api.types.is_datetime64_any_dtype(t[c]):
                    for celda in ws[letra][1:]:
                        celda.number_format = "DD/MM/YYYY"
                    ws.column_dimensions[letra].width = 14
                    continue
                ancho = max(
                    [len(str(c))]
                    + t[c].astype(str).str.len().head(200).tolist()
                )
                ws.column_dimensions[letra].width = (
                    min(max(ancho + 2, 10), 45)
                )
                # Números: $ sin decimales / % con 1 decimal / kg con miles.
                if not pd.api.types.is_numeric_dtype(t[c]):
                    continue
                if pd.api.types.is_bool_dtype(t[c]):
                    continue
                fmt = spec if isinstance(spec, str) else (
                    spec.get(c) if isinstance(spec, dict) else None
                )
                if fmt is None:
                    fmt = _formato_col(c, t[c])
                if isinstance(fmt, (list, tuple)):
                    # Un formato por fila (columna con métricas mezcladas).
                    for celda, f in zip(ws[letra][1:], fmt):
                        if f:
                            celda.number_format = f
                else:
                    for celda in ws[letra][1:]:
                        celda.number_format = fmt
                # Con $ y separadores el texto ocupa más: un poco más de ancho.
                ws.column_dimensions[letra].width = max(
                    ws.column_dimensions[letra].width, 13
                )
    return buf.getvalue()


def boton_excel(nombre, hojas, key, formatos=None):
    """Botón de descarga de un .xlsx con las tablas de la solapa."""
    hojas = {n: t for n, t in hojas.items() if t is not None and len(t)}
    if not hojas:
        return
    st.download_button(
        "Descargar Excel",
        data=_excel_bytes(hojas, formatos),
        file_name=f"{nombre}_{desde:%Y-%m-%d}_a_{hasta:%Y-%m-%d}.xlsx",
        mime=("application/vnd.openxmlformats-officedocument"
              ".spreadsheetml.sheet"),
        key=key,
    )


# ---------------------------------------------------------------------------
# Navegación progresiva (árbol de descomposición)
# ---------------------------------------------------------------------------
# Reemplaza los cruces sueltos de las solapas Proveedores y Productos (SKU):
# un único recorrido por niveles (ej. Línea → Canal → Vendedor → Cliente →
# SKU) donde cada nivel muestra SOLO lo compatible con lo ya elegido y la
# participación % se calcula dentro del nivel anterior.

def _cols_cm(cols):
    """Quita las columnas de CM si el rol no puede verlas."""
    if mostrar_cm:
        return cols
    return [c for c in cols if c not in ("cm", "share_cm", "cm_pct")]


def _barras_share(g, col_dim, etiqueta, col_val, col_share, top_n=12):
    """Barras horizontales de composición: top N + 'OTRAS', con el share
    % como texto. Devuelve la figura lista para st.plotly_chart."""
    top = g.nlargest(top_n, col_val).copy()
    resto = g[~g[col_dim].isin(top[col_dim])]
    if len(resto):
        fila = {col_dim: f"OTRAS ({len(resto)})",
                col_val: resto[col_val].sum(),
                col_share: resto[col_share].sum()}
        top = pd.concat([top, pd.DataFrame([fila])], ignore_index=True)
    top = top.sort_values(col_val)
    fig = px.bar(
        top, x=col_val, y=col_dim, orientation="h",
        text=top[col_share].map(lambda v: f"{v:.1f} %"),
    )
    fig.update_traces(textposition="outside", cliponaxis=False,
                      marker_color="#00b87a")
    fig.update_layout(
        template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=10, r=60, t=10, b=10),
        xaxis_title=None, yaxis_title=None,
        height=max(300, 30 * len(top) + 60),
    )
    return fig


# Métrica principal elegible en el drill: (columna de valor, columna de share)
METRICAS_DRILL = {
    "Facturación": ("subtotalNeto", "share_fc"),
    "Kilos": ("kilos", "share_kg"),
    "Contribución": ("cm", "share_cm"),
}


def render_drill(df_base, niveles, key, root_id=None):
    """Navegación progresiva por niveles (tipo árbol de descomposición).

    df_base : datos ya filtrados (filtros globales y, si aplica, la raíz,
              ej. el producto elegido en el ranking).
    niveles : lista de (etiqueta, columna) en orden de navegación.
              El último nivel es solo informativo (no se puede seleccionar).
    key     : prefijo único para session_state (permite varios drills).
    root_id : identificador de la raíz. Si cambia (ej. se elige otro
              producto), el recorrido se reinicia solo.

    El recorrido vive en session_state como lista de valores elegidos.
    En cada corrida se valida contra los datos vigentes: si un filtro
    global dejó afuera un valor elegido, el camino se corta en el último
    nivel válido (evita pantallas vacías).

    Devuelve la tabla completa del nivel visible (renombrada y sin CM si
    el rol no la ve) para sumarla al Excel de la solapa, o None.
    """
    k_path, k_nonce, k_root = f"{key}_path", f"{key}_nonce", f"{key}_root"
    if k_path not in st.session_state:
        st.session_state[k_path] = []
    if k_nonce not in st.session_state:
        # El nonce entra en la key de la tabla clickeable de cada nivel y se
        # incrementa CADA vez que el recorrido cambia (avanzar, volver,
        # reiniciar). Así el widget nuevo nace sin selección "pegada" de una
        # visita anterior al mismo nivel (gotcha clásico de Streamlit).
        st.session_state[k_nonce] = 0

    def _set_path(nuevo):
        st.session_state[k_path] = nuevo
        st.session_state[k_nonce] += 1

    # Cambió la raíz (ej. otro producto en el ranking): arrancar de cero.
    if st.session_state.get(k_root, "__sin_raiz__") != root_id:
        st.session_state[k_root] = root_id
        _set_path([])

    if df_base.empty:
        st.info("No hay datos para navegar con los filtros actuales.")
        return None

    # --- Validación del recorrido contra los datos vigentes ---------------
    path = st.session_state[k_path]
    d = df_base
    validos = []
    for i, val in enumerate(path):
        if i >= len(niveles) - 1:
            break  # nunca se navega más allá del anteúltimo nivel
        d2 = d[d[niveles[i][1]].astype(str) == str(val)]
        if d2.empty:
            break
        validos.append(val)
        d = d2
    if len(validos) != len(path):
        _set_path(validos)
        path = validos

    nonce = st.session_state[k_nonce]
    nivel = len(path)
    etiqueta_niv, col_niv = niveles[nivel]
    es_hoja = nivel == len(niveles) - 1

    # --- Métrica principal + reiniciar -------------------------------------
    ops_met = ["Facturación", "Kilos"] + (["Contribución"] if mostrar_cm else [])
    c_met, c_res = st.columns([3, 1])
    met = c_met.radio("Métrica principal", ops_met, horizontal=True,
                      key=f"{key}_met")
    col_val, col_share = METRICAS_DRILL[met]
    c_res.markdown("<div style='height:1.7rem'></div>", unsafe_allow_html=True)
    if c_res.button("⟲ Reiniciar recorrido", key=f"{key}_reset",
                    disabled=not path, use_container_width=True):
        _set_path([])
        st.rerun()

    # --- Recorrido (migas): clic en una miga vuelve a ese nivel ------------
    st.caption("Recorrido: " + " → ".join(
        f"**{e}**" if i == nivel else e for i, (e, _) in enumerate(niveles)
    ))
    if path:
        cols_bc = st.columns(max(len(path), 4))
        for i, val in enumerate(path):
            _et = niveles[i][0]
            _txt = str(val) if len(str(val)) <= 22 else str(val)[:21] + "…"
            if cols_bc[i].button(
                f"✕ {_et}: {_txt}", key=f"{key}_bc_{i}",
                use_container_width=True,
                help=f"Quitar este nivel (y los siguientes) y volver a "
                     f"elegir {_et}.",
            ):
                _set_path(path[:i])
                st.rerun()

        # Métricas del contexto acumulado (todo lo seleccionado hasta acá).
        _fc = d["subtotalNeto"].sum()
        _kg = d["kilos"].sum()
        _fc_base = df_base["subtotalNeto"].sum()
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Facturación", fmt_money(_fc))
        m1.caption(f"{(_fc / _fc_base * 100) if _fc_base else 0:.1f} % del total")
        m2.metric("Kilos", fmt_kg(_kg))
        if mostrar_cm:
            m3.metric("Contribución",
                      fmt_money(_fc - d["costo_unitario"].sum()))
        else:
            m3.metric("Precio medio",
                      (fmt_money(_fc / _kg) + " /kg") if _kg else "—")
        m4.metric("Clientes · SKUs",
                  f"{d['idCliente'].nunique()} · {d['idArticulo'].nunique()}")

    # --- Nivel actual -------------------------------------------------------
    g = dp.agrupar_dim(d, col_niv)
    g = g[g[col_niv].astype(str).str.strip() != ""]
    if g.empty:
        st.info(f"No hay {etiqueta_niv.lower()} para esta selección.")
        return None
    # share_cm ya viene de agrupar_dim (mismo criterio que share_fc/share_kg).
    g = g.sort_values(col_val, ascending=False).reset_index(drop=True)

    st.subheader(f"Nivel {nivel + 1} de {len(niveles)} · {etiqueta_niv}")

    # Tabla completa del nivel (va al Excel y a la vista final).
    _cf = [col_niv, "kilos", "share_kg", "subtotalNeto", "share_fc",
           "cm", "share_cm", "cm_pct", "precio_kg", "clientes", "skus"]
    if col_niv == "dsArticulo":
        _cf.remove("skus")  # cada fila ya ES un SKU
    t_full = g[_cols_cm(_cf)].rename(columns={col_niv: etiqueta_niv,
                                              **COLS_DIM})

    if es_hoja:
        # Último nivel: solo se muestra, no se navega más.
        if path:
            st.caption("Resultado filtrado por: " + " · ".join(
                f"{niveles[i][0]} = {v}" for i, v in enumerate(path)
            ))
        st.dataframe(t_full.style.format(FMT_DIM),
                     use_container_width=True, hide_index=True)
    else:
        st.caption(
            f"Hacé clic en una fila para abrir el siguiente nivel "
            f"(→ {niveles[nivel + 1][0]}). Part. % = participación dentro "
            f"de lo ya seleccionado."
        )
        c_graf, c_tab = st.columns([1.15, 1.45])
        with c_graf:
            st.plotly_chart(
                _barras_share(g, col_niv, etiqueta_niv, col_val, col_share),
                use_container_width=True,
                # key explícita: en el nivel 1 esta figura puede ser idéntica
                # al gráfico de composición de arriba y sin key Streamlit
                # colisiona los IDs autogenerados.
                key=f"{key}_fig_{nivel}_{nonce}",
            )
        with c_tab:
            _mn = COLS_DIM[col_val]  # nombre lindo de la métrica elegida
            t_click = g[[col_niv, col_val, col_share, "clientes",
                         "skus"]].rename(columns={
                col_niv: etiqueta_niv, col_val: _mn, col_share: "Part. %",
                "clientes": "Clientes", "skus": "SKUs",
            })
            ev = st.dataframe(
                t_click.style.format({_mn: FMT_DIM.get(_mn),
                                      "Part. %": fmt_pct}),
                use_container_width=True, hide_index=True,
                on_select="rerun", selection_mode="single-row",
                key=f"{key}_sel_{nivel}_{nonce}",
                height=min(420, 36 * (len(t_click) + 1) + 20),
            )
            filas = ev.selection.rows if ev and ev.selection else []
            if filas:
                _set_path(path + [str(g.iloc[filas[0]][col_niv])])
                st.rerun()
        with st.expander(f"Ver tabla completa del nivel ({len(t_full)})"):
            st.dataframe(t_full.style.format(FMT_DIM),
                         use_container_width=True, hide_index=True)

    return t_full


# ---------------------------------------------------------------------------
# Tabs
# ---------------------------------------------------------------------------

# La solapa "Acuerdos McCain" toca el COSTO, así que solo la ve el dueño
# (los supervisores no ven CM). Se arma la lista de tabs según el rol.
_labels_tabs = ["Resumen", "Proveedores", "Canales", "Productos (SKU)",
                "Altas y Bajas", "Vendedores", "Alertas", "Metas"]
if mostrar_cm:
    _labels_tabs.append("Acuerdos McCain")

_tabs = st.tabs(_labels_tabs)
(tab_resumen, tab_lineas, tab_canales, tab_prod, tab_clientes,
 tab_vend, tab_alertas, tab_metas) = _tabs[:8]
tab_acuerdos = _tabs[8] if mostrar_cm else None


# --- TAB RESUMEN ----------------------------------------------------------
with tab_resumen:
    m = dp.metricas_generales(df)

    # --- Proyección a fin de mes ------------------------------------------
    # Run-rate lineal: extrapola lo acumulado hasta hoy al total del mes.
    # El factor NO es único para todos: se calcula por vendedor y se pondera,
    # porque Food Service factura solo dos a cuatro días fijos por semana
    # (ver dp.DIAS_FACTURACION) y proyectarlo contra días hábiles lo
    # distorsiona. Para el resto de los canales cuenta días hábiles: los
    # domingos no se trabaja y los feriados de FERIADOS tampoco.
    # Es el mismo criterio que usa el seguimiento de metas, así que los dos
    # números del tablero cierran entre sí.
    # Solo aplica al mes EN CURSO; los meses cerrados ya están completos.
    #
    # Los feriados salen de dp.FERIADOS (calendario nacional, ver
    # data_pipeline.py) y se aplican SOLOS: no hace falta pasarlos. Antes esta
    # solapa tenía su propia lista con un único feriado y la solapa Metas no
    # pasaba ninguno, así que las dos proyecciones no coincidían.
    factor = 1.0
    proyectar = False
    if es_mes_actual:
        _ini_mes = dt.date(hasta.year, hasta.month, 1)
        _fin_mes = dt.date(hasta.year, hasta.month,
                           calendar.monthrange(hasta.year, hasta.month)[1])
        _ult = df["fechaComprobate"].max()
        _corte_res = _ult.date() if pd.notna(_ult) else hasta
        factor, proyectar = dp.factor_proyeccion_ponderado(
            df, _ini_mes, _corte_res, _fin_mes)

    def proy(col, valor, fmt, escala=True):
        """Muestra debajo de la métrica la proyección a fin de mes.

        escala=True  -> métrica aditiva (se multiplica por el run-rate).
        escala=False -> métrica de tasa/ratio (se mantiene estable).
        """
        if not proyectar:
            return
        pv = valor * factor if escala else valor
        col.caption(f"Proy. fin de mes: {fmt(pv)}")

    def _int(x):
        return f"{round(x):,}".replace(",", ".")

    # Solo se proyectan las métricas aditivas que se leen como volumen del mes.
    # Las de tasa (CM %, precio medio, ticket) no se pueden extrapolar por run
    # rate —el número proyectado da igual al actual y no aporta nada— y las de
    # conteo (clientes, SKUs) tampoco, porque se repiten entre días en vez de
    # sumarse. Antes se mostraban igual y confundían más de lo que ayudaban.
    _METRICAS_PROYECTADAS = {
        "Facturación neta", "Kilos vendidos", "Contribución marginal",
    }

    # Cada métrica: (etiqueta, valor a mostrar, valor para proyectar,
    # formato o None, escala). El formato se usa para la proyección y para el
    # Excel, por eso se mantiene aunque la métrica no se proyecte. Las de CM
    # solo se agregan para el rol dueño, así supervisores nunca las reciben.
    metricas = [
        ("Facturación neta", fmt_money(m["subtotal_neto"]),
         m["subtotal_neto"], fmt_money, True),
        ("Kilos vendidos", fmt_kg(m["total_kilos"]),
         m["total_kilos"], fmt_kg, True),
    ]
    if mostrar_cm:
        metricas += [
            ("Contribución marginal", fmt_money(m["contribucion_marginal"]),
             m["contribucion_marginal"], fmt_money, True),
            ("CM %", fmt_pct(m["cm_pct"]),
             m["cm_pct"], fmt_pct, False),
        ]
    metricas += [
        ("Precio medio / kg", fmt_money(m["precio_medio_kg"]),
         m["precio_medio_kg"], fmt_money, False),
        ("Clientes únicos", f"{m['n_clientes']:,}".replace(",", "."),
         m["n_clientes"], _int, True),
        ("Ticket promedio", fmt_money(m["ticket_promedio"]),
         m["ticket_promedio"], fmt_money, False),
        ("SKUs vendidos", f"{m['n_skus']:,}".replace(",", "."),
         m["n_skus"], _int, True),
    ]

    # --- Cobertura --------------------------------------------------------
    # Qué porción de la cartera y del surtido del año se tocó en el período.
    # Van sin proyección: el % no se puede extrapolar por regla de tres porque
    # los clientes se repiten, no se suman.
    _cob = dp.cobertura_total(df, df_universo)
    metricas += [
        ("Cobertura de clientes", fmt_pct(_cob["cob_clientes"]),
         _cob["cob_clientes"], None, False),
        ("Cobertura de SKUs", fmt_pct(_cob["cob_skus"]),
         _cob["cob_skus"], None, False),
    ]

    # Tablas que junta esta solapa para el Excel descargable.
    hojas_resumen = {"KPIs": pd.DataFrame(
        [(lbl, pval) for lbl, _disp, pval, _pf, _e in metricas],
        columns=["Métrica", "Valor"],
    )}
    # "Valor" mezcla plata, kilos, % y conteos: el formato va fila por fila,
    # deducido de la función de formato con la que se muestra cada métrica.
    _XL_POR_FMT = {fmt_money: XL_MONEY, fmt_kg: XL_KG, fmt_pct: XL_PCT,
                   _int: XL_INT}
    formatos_resumen = {"KPIs": {"Valor": [
        _XL_POR_FMT.get(_pf, XL_DEC1) for _l, _d, _v, _pf, _e in metricas
    ]}}

    # Render en filas de 4 columnas.
    for i in range(0, len(metricas), 4):
        cols = st.columns(4)
        for col, (lbl, disp, pval, pfmt, escala) in zip(cols, metricas[i:i + 4]):
            col.metric(lbl, disp)
            if pfmt is not None and lbl in _METRICAS_PROYECTADAS:
                proy(col, pval, pfmt, escala)

    st.caption(
        f"{m['n_comprobantes']:,}".replace(",", ".") + " comprobantes  ·  "
        + fmt_kg(m["kg_por_cliente"]) + " por cliente (promedio)"
    )
    st.caption(
        f"Cobertura: se le vendió a {_cob['clientes']} de los "
        f"{_cob['universo_clientes']} clientes y se movieron "
        f"{_cob['skus']} de los {_cob['universo_skus']} SKUs que tuvieron "
        f"movimiento en {dp.ANIO}."
    )

    st.divider()

    # --- Evolución mensual (canal / subcanal / vendedor) --------------------
    # Usa la serie mensual histórica (data/serie_mensual.parquet),
    # INDEPENDIENTE del filtro de período: muestra todos los meses 2025–2026
    # para comparar.
    st.subheader("Evolución mensual")

    if not os.path.exists(dp.SERIE_PATH):
        st.warning(
            "Todavía no existe la serie histórica.\n\n"
            "Generala UNA vez con el backfill:\n\n"
            "```\npython backfill_serie.py\n```\n\n"
            "Después el pipeline normal la mantiene actualizada sola."
        )
    else:
        serie = cargar_serie(
            os.path.getmtime(dp.SERIE_PATH),
            os.path.getmtime(PARQUET_PATH),
            _mtime_acuerdos(),
        )

        if serie.empty:
            st.info("La serie histórica está vacía.")
        else:
            METRICAS_EVOL = {
                "Facturación neta": ("subtotalNeto", fmt_money),
                "Kilos": ("kilos", fmt_kg),
            }
            # Supervisores no ven las métricas de CM en el selector.
            if mostrar_cm:
                METRICAS_EVOL["Contribución marginal (MB $)"] = ("cm", fmt_money)
                METRICAS_EVOL["CM % (margen)"] = ("cm_pct", fmt_pct)
            METRICAS_EVOL["Precio medio $/kg"] = ("precio_kg", fmt_money)

            c1, c2, c3 = st.columns([1.5, 0.9, 1.2])
            nombre_metrica = c1.selectbox(
                "Métrica", list(METRICAS_EVOL.keys()), index=0
            )
            nivel = c2.radio(
                "Abrir por", ["Canal", "Subcanal", "Vendedor"], horizontal=True
            )
            moneda = c3.radio(
                "Moneda", ["Corriente", "Constante (s/ inflación)"],
                horizontal=True,
                help="Corriente = pesos de cada mes (nominal). "
                     "Constante = todo llevado a pesos de hoy con el IPC del "
                     "INDEC, para comparar sin el efecto de la inflación.",
            )
            dim = {
                "Canal": "dsCanalMkt",
                "Subcanal": "dsSubcanalMKT",
                "Vendedor": "dsVendedor",
            }[nivel]

            # Respeta los filtros globales de canal/subcanal/vendedor si están
            # activos (las sumas crudas se re-agregan bien sea cual sea el
            # nivel elegido para abrir el gráfico).
            s = serie.copy()
            if seleccion.get("dsCanalMkt"):
                s = s[s["dsCanalMkt"].astype(str).str.strip()
                      .isin(seleccion["dsCanalMkt"])]
            if seleccion.get("dsSubcanalMKT"):
                s = s[s["dsSubcanalMKT"].astype(str).str.strip()
                      .isin(seleccion["dsSubcanalMKT"])]
            if seleccion.get("dsVendedor"):
                s = s[s["dsVendedor"].astype(str).str.strip()
                      .isin(seleccion["dsVendedor"])]

            # Re-agrega al nivel elegido (las sumas crudas se re-agregan bien).
            g = (
                s.groupby(["anio_mes", dim], dropna=False)
                .agg(
                    kilos=("kilos", "sum"),
                    subtotalNeto=("subtotalNeto", "sum"),
                    cm=("cm", "sum"),
                )
                .reset_index()
            )

            # --- Pesos constantes: deflactar $ con el IPC del INDEC ---------
            base_mes = None
            nota_moneda = "Pesos corrientes (nominales, de cada mes)."
            if moneda.startswith("Constante"):
                _ipc_mtime = (os.path.getmtime(dp.IPC_PATH)
                              if os.path.exists(dp.IPC_PATH) else None)
                ipc = cargar_ipc(_ipc_mtime)
                factores, base_mes = dp.factores_constantes(ipc)
                if not factores:
                    st.warning(
                        "No hay IPC disponible todavía (corré el pipeline o "
                        "esperá a que INDEC responda). Mostrando pesos corrientes."
                    )
                else:
                    # Factor por mes; meses sin IPC (ej. mes en curso) usan el
                    # último factor disponible (≈1 respecto del mes base).
                    ult = min(factores.values())  # el del mes más reciente
                    fac = g["anio_mes"].map(factores).fillna(ult)
                    g["subtotalNeto"] = g["subtotalNeto"] * fac
                    g["cm"] = g["cm"] * fac
                    nota_moneda = (
                        f"Pesos constantes de {base_mes} (deflactado con IPC "
                        f"Nivel General INDEC). Kilos y % no se ven afectados."
                    )

            # Métricas derivadas (porcentaje y $/kg se calculan acá, no se guardan).
            den_fc = g["subtotalNeto"].replace(0, pd.NA)
            den_kg = g["kilos"].replace(0, pd.NA)
            g["cm_pct"] = (g["cm"] / den_fc * 100).fillna(0)
            g["precio_kg"] = (g["subtotalNeto"] / den_kg).fillna(0)

            col_val, _fmt = METRICAS_EVOL[nombre_metrica]
            g = g.sort_values(["anio_mes", dim])

            # --- Total por mes (suma de todos los canales/subcanales) --------
            # Las métricas aditivas se suman; los % y $/kg se recalculan sobre
            # los totales para que el "Total" sea correcto (no un promedio).
            tot = (
                g.groupby("anio_mes", as_index=False)
                .agg(kilos=("kilos", "sum"),
                     subtotalNeto=("subtotalNeto", "sum"),
                     cm=("cm", "sum"))
            )
            tot_den_fc = tot["subtotalNeto"].replace(0, pd.NA)
            tot_den_kg = tot["kilos"].replace(0, pd.NA)
            tot["cm_pct"] = (tot["cm"] / tot_den_fc * 100).fillna(0)
            tot["precio_kg"] = (tot["subtotalNeto"] / tot_den_kg).fillna(0)
            tot = tot.sort_values("anio_mes")

            fig = px.line(
                g, x="anio_mes", y=col_val, color=dim, markers=True,
            )
            # La línea de "Total" solo suma valor cuando se abre por Canal
            # (pocas categorías). En Subcanal/Vendedor hay demasiadas líneas
            # y el total se pisa con ellas, así que se omite.
            if nivel == "Canal":
                fig.add_scatter(
                    x=tot["anio_mes"], y=tot[col_val], mode="lines+markers",
                    name="Total", line=dict(color="#e5e7eb", width=3, dash="dash"),
                    marker=dict(size=6),
                )
            fig.update_layout(
                template="plotly_dark",
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                margin=dict(l=10, r=10, t=10, b=10),
                legend=dict(title=nivel, orientation="h", y=-0.2),
                xaxis_title="Mes",
                yaxis_title=nombre_metrica,
                height=440,
            )
            st.plotly_chart(fig, use_container_width=True)
            st.caption(
                f"{nota_moneda}  ·  Serie completa (no depende del filtro de "
                "período de arriba). El mes en curso puede estar incompleto."
            )

            # Tabla pivote opcional (meses en columnas) para ver los números.
            with st.expander("Ver tabla de valores"):
                piv = g.pivot_table(
                    index=dim, columns="anio_mes", values=col_val,
                    aggfunc="sum",
                )
                st.dataframe(
                    piv.style.format(_fmt), use_container_width=True
                )
                hojas_resumen["Evolución mensual"] = piv.reset_index()
                # Las columnas de la pivote son meses ("2026-01"): el formato
                # no se puede deducir del nombre, lo define la métrica elegida.
                fmt_evol = {
                    fmt_money: XL_MONEY, fmt_kg: XL_KG, fmt_pct: XL_PCT,
                }.get(_fmt, XL_DEC1)
                formatos_resumen["Evolución mensual"] = fmt_evol

    st.divider()
    boton_excel("resumen", hojas_resumen, key="xlsx_resumen",
                formatos=formatos_resumen)


# --- TAB LÍNEAS (gestión comercial por línea / marca) ----------------------
# Estructura en 3 niveles: panorama (qué pesa cada línea), apertura de una
# línea (por vendedor / canal / canal→vendedor / canal→subcanal) y la mirada
# inversa vendedor → línea → producto. Respeta los filtros globales de arriba:
# trabaja sobre el mismo `df` ya filtrado, así los shares se recalculan sobre
# la selección vigente.
with tab_lineas:
    # Línea "estricta": lookup por artículo; SKUs sin regla -> SIN ASIGNAR
    # (solo en esta solapa; en el resto siguen cayendo al proveedor).
    dfl = dp.agregar_linea_estricta(df)
    g_lin = dp.agrupar_dim(dfl, "linea_producto")
    hojas_lineas = {}  # tablas para el Excel descargable de la solapa

    # --- 1) Panorama: cuánto pesa cada línea --------------------------------
    st.subheader("Composición de la venta por línea de producto")

    _total_fc = g_lin["subtotalNeto"].sum()
    _fc_sin = g_lin.loc[
        g_lin["linea_producto"] == dp.SIN_ASIGNAR, "subtotalNeto"
    ].sum()
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Líneas activas", f"{len(g_lin)}")
    c1.caption("con ventas en el período/filtros")
    c2.metric("Línea principal", g_lin.iloc[0]["linea_producto"])
    c2.caption(f"{g_lin.iloc[0]['share_fc']:.1f} % de la facturación")
    c3.metric("Top 3 líneas", fmt_pct(g_lin["share_fc"].head(3).sum()))
    c3.caption("de la facturación (concentración)")
    c4.metric("Sin asignar", fmt_money(_fc_sin))
    c4.caption(
        f"{(_fc_sin / _total_fc * 100) if _total_fc else 0:.1f} % de la "
        "facturación en SKUs sin línea"
    )

    met_lin = st.radio(
        "Ver composición por", ["Facturación", "Kilos"],
        horizontal=True, key="lin_met",
    )
    _cv = "subtotalNeto" if met_lin == "Facturación" else "kilos"
    _cs = "share_fc" if met_lin == "Facturación" else "share_kg"
    st.plotly_chart(
        _barras_share(g_lin, "linea_producto", "Línea", _cv, _cs),
        use_container_width=True,
    )

    with st.expander("Ver tabla completa de líneas"):
        _cols = _cols_cm(["linea_producto", "kilos", "share_kg", "subtotalNeto",
                          "share_fc", "cm", "share_cm", "cm_pct", "precio_kg",
                          "clientes", "skus"])
        t_lineas = g_lin[_cols].rename(
            columns={"linea_producto": "Línea", **COLS_DIM})
        st.dataframe(
            t_lineas.style.format(FMT_DIM),
            use_container_width=True, hide_index=True,
        )
        hojas_lineas["Líneas"] = t_lineas

    # --- 2) Navegación progresiva: Línea → Canal → ... → SKU -----------------
    st.divider()
    st.subheader("Navegación por línea")
    st.caption(
        "Elegí una línea y bajá nivel por nivel: Línea → Canal → Vendedor → "
        "Cliente → SKU. Cada nivel muestra solo lo compatible con lo ya "
        "seleccionado; con las etiquetas del recorrido volvés a un nivel "
        "anterior."
    )
    t_drill_lin = render_drill(
        dfl,
        [("Línea", "linea_producto"), ("Canal", "dsCanalMkt"),
         ("Vendedor", "dsVendedor"), ("Cliente", "nombreCliente"),
         ("SKU", "dsArticulo")],
        key="drill_prov",
    )
    if t_drill_lin is not None:
        hojas_lineas["Navegación"] = t_drill_lin

    # --- 3) SKUs sin línea asignada ------------------------------------------
    st.divider()
    _sin = dfl[dfl["linea_producto"] == dp.SIN_ASIGNAR]
    _n_sin = _sin["dsArticulo"].nunique()
    with st.expander(f"SKUs sin línea asignada ({_n_sin})"):
        if _sin.empty:
            st.success("Todos los SKUs del período tienen línea asignada.")
        else:
            g_sin = dp.agrupar_multi(_sin, ["dsArticulo", "proveedor"])
            _cols_s = ["dsArticulo", "proveedor", "kilos", "subtotalNeto",
                       "clientes"]
            t_sin = g_sin[_cols_s].rename(columns={
                "dsArticulo": "Producto", "proveedor": "Proveedor",
                **COLS_DIM,
            })
            st.dataframe(
                t_sin.style.format(FMT_DIM),
                use_container_width=True, hide_index=True,
            )
            hojas_lineas["SKUs sin línea"] = t_sin
            st.caption(
                "Para clasificarlos: agregar la fila correspondiente en "
                "data/proveedor_objetivo_lookup.csv (columnas idArticulo,"
                "marca_linea) con el código del artículo. El próximo run del "
                "pipeline los toma solo."
            )

    st.divider()
    boton_excel("proveedores", hojas_lineas, key="xlsx_lineas")


# --- TAB CANALES ----------------------------------------------------------
with tab_canales:
    st.subheader("Detalle por canal")
    t_canales = tabla_dim(
        dp.agregar_cobertura(dp.por_canal(df), "dsCanalMkt", df_universo),
        "Canal", "dsCanalMkt", mostrar_skus=True)

    # Para dueños: torta de share + CM % por canal lado a lado.
    # Para supervisores: solo la torta (a ancho completo), sin CM %.
    col_a, col_b = st.columns(2) if mostrar_cm else (st.container(), None)
    with col_a:
        st.caption("Share de facturación por canal")
        _pc = dp.por_canal(df)
        fig_torta = px.pie(
            _pc, names="dsCanalMkt", values="share_fc", hole=0.4,
        )
        fig_torta.update_traces(textposition="inside", textinfo="percent+label")
        fig_torta.update_layout(
            template="plotly_dark",
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            margin=dict(l=10, r=10, t=10, b=10),
            showlegend=False,
            height=360,
        )
        st.plotly_chart(fig_torta, use_container_width=True)
    if mostrar_cm:
        with col_b:
            st.caption("CM % por canal")
            st.bar_chart(dp.por_canal(df).set_index("dsCanalMkt")["cm_pct"])

    st.divider()
    st.subheader("Detalle por subcanal")
    t_subcanales = tabla_dim(
        dp.agregar_cobertura(dp.por_subcanal(df), "dsSubcanalMKT", df_universo),
        "Subcanal", "dsSubcanalMKT", mostrar_skus=True)

    st.subheader("Detalle por marca / línea")
    # Acá la cobertura se lee como penetración de la marca: de los clientes
    # que compraron esa marca alguna vez en el año, cuántos la compraron en
    # el mes.
    t_marcas = tabla_dim(
        dp.agregar_cobertura(dp.por_proveedor(df), "marca_linea", df_universo),
        "Marca / Línea", "marca_linea")

    st.divider()
    boton_excel("canales", {
        "Canales": t_canales,
        "Subcanales": t_subcanales,
        "Marca - Línea": t_marcas,
    }, key="xlsx_canales")


# --- TAB PRODUCTOS (SKU) --------------------------------------------------
with tab_prod:
    prod = dp.ranking_productos(df)
    hojas_prod = {}  # tablas para el Excel descargable de la solapa

    if prod.empty:
        st.info("No hay productos en el período seleccionado.")
    else:
        n_a = int((prod["ABC"] == "A").sum())
        n_b = int((prod["ABC"] == "B").sum())
        n_c = int((prod["ABC"] == "C").sum())
        c1, c2, c3 = st.columns(3)
        c1.metric("SKUs clase A (80% FC)", n_a)
        c2.metric("SKUs clase B (80-95%)", n_b)
        c3.metric("SKUs clase C (resto)", n_c)

        st.divider()
        # Filtro por clase ABC: si elegís "A" (o varias), el ranking de abajo
        # muestra el top SOLO de esa/esas clases. Vacío = todas. Sin esto,
        # el Top N siempre caía en clase A porque "prod" viene ordenado por
        # facturación descendente (los primeros N son casi siempre A).
        c_buscar, c_abc, c_top = st.columns([3, 1.4, 1.4])
        with c_buscar:
            buscar = st.text_input(
                "Buscar producto por nombre",
                placeholder="Buscar Producto",
                key="buscar_prod",
            )
        with c_abc:
            abc_uno = st.pills(
                "Clase ABC", ["A", "B", "C"], selection_mode="single",
                default=None, key="abc_sel_prod",
            )
        with c_top:
            top_n = st.select_slider(
                "Top N", options=[5, 10, 15, 25, 50], value=10,
                key="top_n_prod",
            )
        abc_sel = [abc_uno] if abc_uno else []
        prod_f = prod[prod["ABC"].isin(abc_sel)] if abc_sel else prod

        # Búsqueda por nombre: filtra el ranking por coincidencia parcial
        # (sin distinguir mayúsculas/acentos) en el nombre del producto.
        if buscar and buscar.strip():
            termino = buscar.strip()
            prod_f = prod_f[
                prod_f["dsArticulo"].astype(str)
                .str.normalize("NFKD").str.encode("ascii", "ignore").str.decode("ascii")
                .str.contains(
                    termino.encode("ascii", "ignore").decode("ascii"),
                    case=False, na=False,
                )
            ]

        titulo_clase = f" (clase {abc_uno})" if abc_uno else ""
        titulo_buscar = f' · "{buscar.strip()}"' if buscar and buscar.strip() else ""
        st.subheader(
            f"Ranking de productos con clasificación ABC{titulo_clase} · "
            f"Top {top_n}{titulo_buscar}"
        )
        if prod_f.empty:
            st.info("No hay productos en la clase seleccionada.")
        else:
            cols = ["dsArticulo", "ABC", "kilos", "subtotalNeto", "share_fc",
                    "cm", "share_cm", "cm_pct", "precio_kg", "clientes"]
            # Supervisores no ven Contribución ni CM % en el ranking de SKUs.
            if not mostrar_cm:
                cols = [c for c in cols if c not in ("cm", "share_cm", "cm_pct")]
            # Copia sin renombrar: sirve para recuperar el nombre real del
            # producto a partir de la fila que el usuario seleccione (la
            # selección devuelve la posición de la fila en este mismo orden).
            prod_top = prod_f[cols].head(top_n).reset_index(drop=True)
            # "clientes" = a cuántos clientes distintos se le vende el
            # producto. Antes esta columna se mostraba como "Cobertura", pero
            # es un CONTEO, no un porcentaje: chocaba con la cobertura real
            # (Cob. clientes % / Cob. SKUs %) de Canales y Vendedores. Se
            # renombró a "Clientes que lo compran" para sacar la ambigüedad.
            _REN_PROD = {"dsArticulo": "Producto", **COLS_DIM,
                         "clientes": "Clientes que lo compran"}
            t = prod_top.rename(columns=_REN_PROD)
            # Al Excel va el ranking COMPLETO filtrado (no solo el Top N).
            hojas_prod["Ranking ABC"] = prod_f[cols].rename(columns=_REN_PROD)
            sel_evt = st.dataframe(
                t.style.format(FMT_DIM), use_container_width=True,
                hide_index=True, on_select="rerun",
                selection_mode="single-row", key="tabla_prod",
            )
            st.caption(
                "Hacé clic en un producto para ver su apertura por canal, "
                "vendedor o cliente."
            )

            # --- Apertura del producto seleccionado --------------------------
            # Mismo formato que "Apertura de una línea" en la solapa Avance:
            # métricas del producto + apertura por dimensión a elección
            # (canal, vendedor, cliente o canal → cliente).
            filas_sel = sel_evt.selection.rows if sel_evt and sel_evt.selection else []
            if filas_sel:
                fila = filas_sel[0]
                nombre_prod = prod_top.iloc[fila]["dsArticulo"]
                det = df[df["dsArticulo"].astype(str) == str(nombre_prod)]
                if det.empty:
                    st.info("Sin datos para este producto.")
                else:
                    st.divider()
                    st.subheader(f"Navegación del producto · {nombre_prod}")

                    m_prod = prod_top.iloc[fila]
                    k1, k2, k3, k4 = st.columns(4)
                    k1.metric("Facturación", fmt_money(m_prod["subtotalNeto"]))
                    k1.caption(f"{m_prod['share_fc']:.1f} % del total")
                    k2.metric("Kilos", fmt_kg(m_prod["kilos"]))
                    k3.metric("Precio medio",
                              fmt_money(m_prod["precio_kg"]) + " /kg")
                    k4.metric("Lo compran",
                              f"{int(m_prod['clientes'])} clientes")

                    st.caption(
                        "Bajá nivel por nivel: Canal → Vendedor → Cliente. "
                        "Al final ves qué clientes compraron este producto "
                        "con todos los filtros acumulados. Part. % = "
                        "participación dentro del producto y de lo ya "
                        "seleccionado."
                    )
                    # El recorrido se reinicia solo si se elige otro producto
                    # en el ranking (root_id).
                    t_drill_p = render_drill(
                        det,
                        [("Canal", "dsCanalMkt"),
                         ("Vendedor", "dsVendedor"),
                         ("Cliente", "nombreCliente")],
                        key="drill_prod", root_id=str(nombre_prod),
                    )
                    if t_drill_p is not None:
                        hojas_prod["Navegación producto"] = t_drill_p

    st.divider()
    boton_excel("productos", hojas_prod, key="xlsx_prod")


# --- TAB ALTAS Y BAJAS ----------------------------------------------------
# NOTA: el bloque de RFM (segmentos + top clientes por facturación/frecuencia)
# quedó COMENTADO a pedido, para que la solapa muestre solo altas y bajas.
# No se borró nada: las funciones dp.rfm() y dp.resumen_segmentos() siguen
# vivas en data_pipeline.py, así que para reactivarlo alcanza con descomentar
# el bloque de abajo (y volver a poner "Clientes (RFM)" en _labels_tabs).
with tab_clientes:
    hojas_cli = {}  # tablas para el Excel descargable de la solapa

    # ----- INICIO BLOQUE RFM COMENTADO --------------------------------------
    # r = dp.rfm(df)
    #
    # if r.empty:
    #     st.info("No hay datos suficientes para el RFM.")
    # else:
    #     st.subheader("Segmentos de clientes (RFM)")
    #     seg = dp.resumen_segmentos(r)
    #     t_seg = seg.rename(columns={
    #         "segmento": "Segmento", "clientes": "Clientes",
    #         "facturacion": "Facturación",
    #     })
    #     st.dataframe(
    #         t_seg.style.format({"Facturación": fmt_money}),
    #         use_container_width=True, hide_index=True,
    #     )
    #     hojas_cli["Segmentos"] = t_seg
    #
    #     st.divider()
    #     # Filtro por segmento: si elegís "Campeones" (o varios), las tablas de
    #     # abajo muestran el top SOLO de ese/esos segmento(s). Vacío = todos.
    #     ORDEN_SEG = ["Campeones", "Leales", "Nuevos / Prometedores",
    #                  "En riesgo", "Hibernando / Perdidos"]
    #     segs_disp = [s for s in ORDEN_SEG if s in set(r["segmento"])]
    #     c_seg, c_top = st.columns([3, 1])
    #     seg_sel = c_seg.multiselect(
    #         "Segmento", segs_disp, default=[],
    #         placeholder="Todos los segmentos", key="seg_rfm",
    #     )
    #     top_n = c_top.select_slider(
    #         "Top N", options=[5, 10, 15, 25, 50], value=10, key="top_n_rfm"
    #     )
    #     r_f = r[r["segmento"].isin(seg_sel)] if seg_sel else r
    #     # Al Excel va la lista COMPLETA de clientes del filtro (no el Top N).
    #     if not r_f.empty:
    #         hojas_cli["Clientes RFM"] = (
    #             r_f.sort_values("monetario", ascending=False)
    #             [["nombreCliente", "segmento", "monetario", "frecuencia",
    #               "recencia"]]
    #             .rename(columns={
    #                 "nombreCliente": "Cliente", "segmento": "Segmento",
    #                 "monetario": "Facturación", "frecuencia": "Frecuencia",
    #                 "recencia": "Recencia (días)",
    #             })
    #         )
    #     if r_f.empty:
    #         st.info("No hay clientes en el segmento seleccionado.")
    #     else:
    #         col_a, col_b = st.columns(2)
    #         with col_a:
    #             st.subheader("Top clientes por facturación")
    #             st.dataframe(
    #                 r_f.sort_values("monetario", ascending=False).head(top_n)
    #                 [["nombreCliente", "segmento", "monetario", "frecuencia", "recencia"]]
    #                 .rename(columns={
    #                     "nombreCliente": "Cliente", "segmento": "Segmento",
    #                     "monetario": "Facturación", "frecuencia": "Frecuencia",
    #                     "recencia": "Recencia (días)",
    #                 })
    #                 .style.format({"Facturación": fmt_money}),
    #                 use_container_width=True, hide_index=True,
    #             )
    #         with col_b:
    #             st.subheader("Top clientes por frecuencia")
    #             st.dataframe(
    #                 r_f.sort_values("frecuencia", ascending=False).head(top_n)
    #                 [["nombreCliente", "segmento", "frecuencia", "monetario", "recencia"]]
    #                 .rename(columns={
    #                     "nombreCliente": "Cliente", "segmento": "Segmento",
    #                     "frecuencia": "Frecuencia", "monetario": "Facturación",
    #                     "recencia": "Recencia (días)",
    #                 })
    #                 .style.format({"Facturación": fmt_money}),
    #                 use_container_width=True, hide_index=True,
    #             )
    # ----- FIN BLOQUE RFM COMENTADO -----------------------------------------

    # --- Altas y bajas de clientes ------------------------------------------
    # Compara el MES ELEGIDO arriba contra SU mes anterior (necesita ver los
    # dos meses a la vez, por eso usa el parquet completo y no df_periodo).
    # Los filtros de dimensión (canal, vendedor, etc.) sí aplican.
    st.subheader("Altas y bajas de clientes")

    # Reusa `df_anio` (el detalle completo que se guardó en la línea 447, antes
    # de que `df` se pise con el período elegido). NO volver a llamar a
    # cargar_datos_local() acá: st.cache_data devuelve una COPIA nueva en cada
    # llamada (deserializa lo que tiene guardado), así que cada llamada de más
    # son ~200 MB extra de RAM en CADA re-ejecución de la app. Eso era lo que
    # hacía que el proceso se quedara sin memoria y Streamlit lo matara.
    _df_ab = df_anio
    for _c_ab, _v_ab in seleccion.items():
        if _v_ab:
            _df_ab = _df_ab[_df_ab[_c_ab].astype(str).str.strip().isin(_v_ab)]

    _mes_ant_ab = desde - dt.timedelta(days=1)  # último día del mes anterior
    _f_ab = _df_ab["fechaComprobate"]
    _hay_ant = (
        (_f_ab >= pd.Timestamp(_mes_ant_ab.replace(day=1)))
        & (_f_ab < pd.Timestamp(desde))
    ).any()

    if not _hay_ant:
        st.info(
            f"No hay datos de {_mes_ant_ab:%m/%Y} en el parquet, así que no "
            f"se puede comparar {hasta:%m/%Y} contra su mes anterior."
        )
    else:
        # hoy=hasta: para el mes en curso corta en hoy; para un mes cerrado
        # usa el mes completo. Así funciona con cualquier mes de 2026.
        altas, bajas = dp.altas_bajas(_df_ab, hoy=hasta)

        _nota_curso = (
            f" (al {hasta:%d/%m/%Y}, puede revertirse si compran antes de "
            f"fin de mes)" if es_mes_actual else ""
        )
        st.caption(
            f"Altas: compraron en {hasta:%m/%Y} y no en {_mes_ant_ab:%m/%Y}. "
            f"Bajas: compraron en {_mes_ant_ab:%m/%Y} y no en "
            f"{hasta:%m/%Y}{_nota_curso}."
        )

        _cols_ab = ["nombreCliente", "dsCanalMkt", "dsVendedor", "compras",
                    "kilos", "facturacion", "ultima_compra"]
        _ren_ab = {
            "nombreCliente": "Cliente", "dsCanalMkt": "Canal",
            "dsVendedor": "Vendedor", "compras": "Compras", "kilos": "Kilos",
            "facturacion": "Facturación", "ultima_compra": "Última compra",
        }
        _fmt_ab = {
            "Kilos": fmt_kg, "Facturación": fmt_money,
            "Última compra": lambda x: f"{x:%d/%m/%Y}",
        }

        # --- Corte por canal / vendedor -------------------------------------
        # Un mismo cliente puede facturar por más de un canal o vendedor: acá
        # se lo cuenta en el DOMINANTE del mes (el de mayor facturación), para
        # que la suma de las filas cierre con el total de altas y bajas.
        _dim_ab = st.radio(
            "Abrir altas y bajas por", ["Canal", "Vendedor"],
            horizontal=True, key="ab_dim",
        )
        _col_dim_ab = {"Canal": "dsCanalMkt", "Vendedor": "dsVendedor"}[_dim_ab]

        def _resumen_ab(d, etiqueta):
            """Cantidad de clientes y facturación por canal/vendedor."""
            if d.empty:
                return pd.DataFrame(columns=[_dim_ab, etiqueta,
                                             f"Facturación {etiqueta.lower()}"])
            g = (d.groupby(_col_dim_ab)
                   .agg(**{etiqueta: ("idCliente", "nunique"),
                           f"Facturación {etiqueta.lower()}": ("facturacion", "sum")})
                   .reset_index()
                   .rename(columns={_col_dim_ab: _dim_ab}))
            return g

        _res_alt = _resumen_ab(altas, "Altas")
        _res_baj = _resumen_ab(bajas, "Bajas")
        t_resumen_ab = (
            _res_alt.merge(_res_baj, on=_dim_ab, how="outer")
            .fillna({"Altas": 0, "Bajas": 0,
                     "Facturación altas": 0, "Facturación bajas": 0})
        )
        if not t_resumen_ab.empty:
            t_resumen_ab["Altas"] = t_resumen_ab["Altas"].astype(int)
            t_resumen_ab["Bajas"] = t_resumen_ab["Bajas"].astype(int)
            t_resumen_ab["Neto"] = t_resumen_ab["Altas"] - t_resumen_ab["Bajas"]
            t_resumen_ab = t_resumen_ab[
                [_dim_ab, "Altas", "Bajas", "Neto",
                 "Facturación altas", "Facturación bajas"]
            ].sort_values("Neto").reset_index(drop=True)
            st.dataframe(
                t_resumen_ab.style.format({
                    "Facturación altas": fmt_money,
                    "Facturación bajas": fmt_money,
                }),
                use_container_width=True, hide_index=True,
            )
            hojas_cli[f"Altas y bajas por {_dim_ab.lower()}"] = t_resumen_ab
            st.caption(
                f"Neto = altas − bajas. Ordenado de peor a mejor {_dim_ab.lower()}."
            )

        # Filtro opcional: recorta el detalle de abajo al canal/vendedor elegido.
        _vals_dim_ab = sorted(
            set(altas[_col_dim_ab].astype(str)) | set(bajas[_col_dim_ab].astype(str))
        )
        _sel_dim_ab = st.multiselect(
            _dim_ab, _vals_dim_ab, default=[],
            placeholder=f"Todos los {_dim_ab.lower()}es", key="ab_dim_sel",
        )
        if _sel_dim_ab:
            altas = altas[altas[_col_dim_ab].astype(str).isin(_sel_dim_ab)]
            bajas = bajas[bajas[_col_dim_ab].astype(str).isin(_sel_dim_ab)]

        st.divider()

        col_alta, col_baja = st.columns(2)
        with col_alta:
            st.metric("Altas", len(altas))
            if altas.empty:
                st.info("Sin altas en el mes seleccionado.")
            else:
                t_altas = altas[_cols_ab].rename(columns=_ren_ab)
                st.dataframe(
                    t_altas.style.format(_fmt_ab),
                    use_container_width=True, hide_index=True,
                )
                hojas_cli["Altas"] = t_altas
        with col_baja:
            st.metric("Bajas", len(bajas))
            if bajas.empty:
                st.info("Sin bajas: todos los clientes del mes anterior "
                        "volvieron a comprar.")
            else:
                t_bajas = bajas[_cols_ab].rename(columns=_ren_ab)
                st.dataframe(
                    t_bajas.style.format(_fmt_ab),
                    use_container_width=True, hide_index=True,
                )
                hojas_cli["Bajas"] = t_bajas
        st.caption(
            "Las cifras de cada tabla corresponden al mes en que el cliente "
            "compró (altas: mes seleccionado · bajas: su mes anterior)."
        )

    st.divider()
    boton_excel("altas_bajas", hojas_cli, key="xlsx_clientes")


# --- TAB VENDEDORES -------------------------------------------------------
with tab_vend:
    st.subheader("Detalle por vendedor")
    _g_vend = dp.agregar_cobertura(dp.por_vendedor(df), "dsVendedor",
                                   df_universo)
    t_vend = tabla_dim(_g_vend, "Vendedor", "dsVendedor",
                       mostrar_skus=True, mostrar_skus_cliente=True)

    # --- Cobertura de cartera por vendedor ---------------------------------
    # El gráfico ordena de menor a mayor: arriba de todo queda el que dejó
    # más clientes sin visitar, que es la lectura accionable de la reunión.
    if "cob_clientes" in _g_vend.columns and _g_vend["cob_clientes"].notna().any():
        st.divider()
        st.caption("Cobertura de cartera por vendedor (% de sus clientes del "
                   "año a los que le vendió en el período)")
        _cob_v = (
            _g_vend.dropna(subset=["cob_clientes"])
            .sort_values("cob_clientes")
            .set_index("dsVendedor")["cob_clientes"]
        )
        st.bar_chart(_cob_v)
        _flojos = _cob_v[_cob_v < 50]
        if not _flojos.empty:
            st.caption(
                "Debajo del 50 %: "
                + "  ·  ".join(f"{v} ({p:.0f} %)" for v, p in _flojos.items())
                + ".  Revisá si son carteras reales o vendedores dados de "
                  "baja que siguen con clientes asignados."
            )

    st.divider()
    st.caption("Facturación por vendedor")
    st.bar_chart(dp.por_vendedor(df).set_index("dsVendedor")["subtotalNeto"])

    st.divider()
    boton_excel("vendedores", {"Vendedores": t_vend}, key="xlsx_vend")


# --- TAB ALERTAS ----------------------------------------------------------
# Seis preguntas fijas por canal, siempre las mismas, para leer la reunión de
# mesa chica sin tener que armar el Excel a mano.
#
# Antes esta solapa mostraba además los avisos por umbral de dp.alertas()
# (margen negativo, concentración del top 10, Pareto de SKUs). Se sacaron a
# pedido: la mesa chica se lee con las tarjetas y los avisos sueltos arriba
# distraían. La función sigue en data_pipeline.py por si se reactivan.
#
# Usa el parquet COMPLETO y no `df`, porque las lecturas de "caído" necesitan
# ver también el mes anterior. Los filtros de dimensión de arriba sí se
# aplican (mismo criterio que Altas y bajas de clientes).
with tab_alertas:
    st.subheader("Lecturas por canal")

    _df_ins = df_anio   # ver nota en "Altas y bajas": reusar, no recargar
    for _c_ins, _v_ins in seleccion.items():
        if _v_ins:
            _df_ins = _df_ins[_df_ins[_c_ins].astype(str).str.strip().isin(_v_ins)]

    _ini_prev, _fin_prev = dp.ventana_anterior(desde, hasta)
    _hab_act = dp.dias_habiles(desde, hasta)
    _hab_prev = dp.dias_habiles(_ini_prev, _fin_prev)

    st.caption(
        f"Comparación mes vs mes **a igual día**: "
        f"{desde:%d/%m} → {hasta:%d/%m} contra {_ini_prev:%d/%m} → "
        f"{_fin_prev:%d/%m}  ·  {_hab_act} vs {_hab_prev} días hábiles  ·  "
        f"base: facturación neta (sin IVA)."
    )
    # Antes acá salía un st.warning cuando los dos tramos no tenían la misma
    # cantidad de días hábiles. Se sacó a pedido: el dato ya está en la
    # leyenda de arriba ("10 vs 9 días hábiles") y el cartel amarillo aparecía
    # casi todos los meses, así que ensuciaba la solapa sin agregar nada.

    _ins = dp.insights_mesa_chica(_df_ins, desde, hasta)

    if _ins.empty:
        st.info(
            "No hay datos suficientes para armar las lecturas por canal "
            "(hace falta el mes anterior para las comparaciones)."
        )
    else:
        _canales_ins = _ins["Canal"].unique().tolist()
        _sel_ins = st.radio(
            "Canal", _canales_ins, horizontal=True, key="ins_canal",
            label_visibility="collapsed",
        )
        _bloque = _ins[_ins["Canal"] == _sel_ins]

        # Tres tarjetas por fila: entran las 6 en dos filas sin scroll.
        _tarjetas = _bloque.to_dict("records")
        for _i in range(0, len(_tarjetas), 3):
            _cols_ins = st.columns(3)
            for _col_ins, _t in zip(_cols_ins, _tarjetas[_i:_i + 3]):
                # Cada dato de la leyenda va en su propio renglón: en una sola
                # línea corrida los " · " se pierden y no se distingue dónde
                # termina un dato y empieza el otro.
                _det = "".join(
                    f"<div>{html.escape(_p.strip())}</div>"
                    for _p in str(_t["Detalle"]).split(dp.SEP_DETALLE) if _p.strip()
                )
                _col_ins.markdown(
                    f"<div class='ins-card {_t['nivel']}'>"
                    f"<div class='ins-tit'>{html.escape(_t['Insight'])}</div>"
                    f"<div class='ins-prot'>{html.escape(str(_t['Protagonista']))}</div>"
                    f"<div class='ins-val'>{html.escape(str(_t['Valor']))}</div>"
                    f"<div class='ins-det'>{_det}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

        with st.expander("Ver las lecturas de todos los canales en tabla"):
            st.dataframe(
                _ins.drop(columns="nivel"),
                use_container_width=True, hide_index=True,
            )

    st.divider()
    boton_excel(
        "lecturas_por_canal",
        {"Lecturas por canal": (_ins.drop(columns="nivel")
                                if not _ins.empty else None)},
        key="xlsx_alertas",
    )


# --- TAB METAS ------------------------------------------------------------
# Reemplaza los Excel de "Cierre / Meta" que se armaban a mano cada mes.
#
# ESTRUCTURA (definida con la gestión comercial): se planifica empezando por
# el CANAL, porque cada supervisor responde por el rendimiento de su canal.
# Después se abre por proveedor/línea y se reparte entre vendedores:
#
#     canal  →  proveedor / línea  →  vendedor
#
# El total del canal se carga a mano (no se deriva de la suma) justamente
# para poder validar que la apertura cierre contra él. Ver validar_metas().
#
# Además se distingue el PRESUPUESTO ANUAL (lo que se estimó a principio de
# año, línea base fija) del OBJETIVO MENSUAL (meta de corto plazo, que se
# reajusta según la realidad reciente: si julio vendió de más porque los
# clientes adelantaron compras, agosto baja). El historial de cargas deja ver
# ese reajuste en vez de perder el número original.
#
# OJO: esta solapa usa el dataset COMPLETO del mes de la meta, NO el `df`
# filtrado de arriba. Si el seguimiento respetara los filtros globales, el
# avance se compararía contra un objetivo que sí es del canal entero y los
# porcentajes darían cualquier cosa.
with tab_metas:
    st.subheader("Metas de venta en kilos")

    _df_full_metas = df_anio   # ver nota en "Altas y bajas": reusar, no recargar
    _fecha_full = _df_full_metas["fechaComprobate"]
    _canales_todos = sorted(
        _df_full_metas["dsCanalMkt"].dropna().astype(str).str.strip()
        .replace({"": None}).dropna().unique().tolist()
    )

    _msg_metas = st.session_state.pop("_metas_ok", None)
    if _msg_metas:
        st.success(_msg_metas)

    # --- Mes de la meta ------------------------------------------------
    # Se ofrecen los meses con datos + el mes siguiente al último, para poder
    # dejar cargada la meta del mes que arranca antes de que haya ventas.
    _hoy = dt.date.today()

    def _mes_siguiente(anio_mes):
        a, m = map(int, str(anio_mes).split("-"))
        return f"{a + 1}-01" if m == 12 else f"{a}-{m + 1:02d}"

    def _mes_anterior(anio_mes):
        a, m = map(int, str(anio_mes).split("-"))
        return f"{a - 1}-12" if m == 1 else f"{a}-{m - 1:02d}"

    _meses_meta = list(_meses_disp)
    _prox = _mes_siguiente(_meses_disp[0])
    if _prox not in _meses_meta:
        _meses_meta = [_prox] + _meses_meta

    c_mes, c_can = st.columns(2)
    _idx_mes = _meses_meta.index(mes_sel) if mes_sel in _meses_meta else 0
    mes_meta = c_mes.selectbox(
        "Mes de la meta", _meses_meta, index=_idx_mes,
        format_func=lambda m: etiqueta_mes(m) + (
            " · a cargar" if m not in _meses_disp else ""),
        key="metas_mes",
        help="Podés cargar la meta del mes que viene antes de que empiece.",
    )
    canal_meta = c_can.selectbox(
        "Canal", ["TODOS"] + _canales_todos, key="metas_canal",
        help="Las metas se cargan por canal: es el punto de partida de la "
             "planificación y de quién responde por el número.",
    )
    _canales_vista = _canales_todos if canal_meta == "TODOS" else [canal_meta]

    # --- Días de venta (base de la proyección) --------------------------
    _desde_m, _ = rango_mes(mes_meta, _hoy)
    _ult_dia_m = dt.date(
        int(mes_meta[:4]), int(mes_meta[5:7]),
        calendar.monthrange(int(mes_meta[:4]), int(mes_meta[5:7]))[1],
    )
    # Corte real: última fecha CON datos del mes (no "hoy"), igual que el
    # filtro de período de arriba.
    _mask_mes = (_fecha_full >= pd.Timestamp(_desde_m)) & (
        _fecha_full < pd.Timestamp(_ult_dia_m) + pd.Timedelta(days=1))
    _df_mes_meta = _df_full_metas[_mask_mes].copy()
    _corte = _df_mes_meta["fechaComprobate"].max()
    _corte = _corte.date() if pd.notna(_corte) else _desde_m

    # Días de venta. Se calculan solos y no se muestran como controles: son el
    # motor de la proyección a fin de mes, no algo para configurar.
    #
    # La proyección real se hace VENDEDOR POR VENDEDOR dentro de
    # dp.seguimiento_metas(): Food Service factura solo dos a cuatro días fijos
    # por semana y proyectarlo contra días hábiles lo distorsiona. Los números
    # de acá son el promedio ponderado por kilos de lo que se está mirando, y
    # se usan únicamente para los textos ("día 6 de 13 de venta", ritmo diario
    # necesario). Con canales que facturan todos los días dan exactamente
    # dias_habiles(), igual que antes.
    #
    # Los feriados nacionales (dp.FERIADOS) se descuentan solos, acá y en la
    # proyección: es el mismo calendario que usa la solapa Resumen.
    dias_pas, dias_tot, _dias_mixto = dp.dias_venta_resumen(
        _df_mes_meta, _desde_m, _corte, _ult_dia_m, canales=_canales_vista)

    # Mes anterior (para comparar contra la realidad reciente)
    _mes_prev = _mes_anterior(mes_meta)
    _desde_p, _ = rango_mes(_mes_prev, _hoy)
    _ult_p = dt.date(
        int(_mes_prev[:4]), int(_mes_prev[5:7]),
        calendar.monthrange(int(_mes_prev[:4]), int(_mes_prev[5:7]))[1],
    )
    _df_mes_prev = _df_full_metas[
        (_fecha_full >= pd.Timestamp(_desde_p))
        & (_fecha_full < pd.Timestamp(_ult_p) + pd.Timedelta(days=1))
    ].copy()

    _metas_all = dp.cargar_metas()
    _hist_all = dp.cargar_historial_metas()
    _cerrado = dias_pas >= dias_tot

    st.caption(
        f"Seguimiento al {_corte:%d/%m/%Y} · "
        + ("mes cerrado" if _cerrado
           else f"día {dias_pas} de {dias_tot} de venta")
        + f" · comparación contra {etiqueta_mes(_mes_prev)}. "
        "Esta solapa ignora los filtros de arriba: el objetivo es del canal "
        "completo, así que el avance también."
        + (" En Food Service cada vendedor se proyecta contra sus días de "
           "facturación, no contra días hábiles: los días que se muestran "
           "acá son el promedio ponderado por kilos." if _dias_mixto else "")
    )

    # Aviso, no bloqueo: los vendedores de Food sin días declarados proyectan
    # con días hábiles (lunes a sábado), que es el criterio viejo.
    _sin_dias = dp.vendedores_sin_dias_facturacion(
        _df_mes_meta[_df_mes_meta["dsCanalMkt"].astype(str).str.strip()
                     .isin(_canales_vista)]
        if "dsCanalMkt" in _df_mes_meta.columns else _df_mes_meta)
    if _sin_dias:
        st.caption(
            "🔸 Sin días de facturación declarados (proyectan con días "
            "hábiles): " + ", ".join(_sin_dias) + "."
        )

    _NIVEL_LBL = {"Canal": "canal", "Proveedor / línea": "proveedor",
                  "Vendedor": "vendedor"}
    _COL_NIVEL = {"canal": None, "proveedor": "marca_linea",
                  "vendedor": "dsVendedor"}

    sub_seg, sub_evol, sub_carga, sub_pres = st.tabs(
        ["Seguimiento", "Evolutivo", "Cargar objetivos", "Presupuesto anual"])

    # =====================================================================
    # SEGUIMIENTO
    # =====================================================================
    with sub_seg:
        _nivel_lbl = st.radio(
            "Abrir por", list(_NIVEL_LBL.keys()), horizontal=True,
            key="metas_nivel_seg",
            help="El objetivo del canal es el que manda. Proveedor y vendedor "
                 "son su apertura: los tres deberían cerrar entre sí.",
        )
        _nivel = _NIVEL_LBL[_nivel_lbl]
        _col_niv = _COL_NIVEL[_nivel]

        # El seguimiento del canal se calcula siempre: es la base de los KPIs.
        # El avance real de un canal es el mismo mire uno el nivel que mire,
        # así que los KPI de arriba no cambian al cambiar de apertura.
        _seg_canal = dp.seguimiento_metas(
            _df_mes_meta, _metas_all, mes_meta,
            dias_pasados=dias_pas, dias_totales=dias_tot,
            df_mes_anterior=_df_mes_prev, canales=_canales_vista,
            nivel="canal",
            desde=_desde_m, corte=_corte, hasta=_ult_dia_m,
        )
        _seg = _seg_canal if _nivel == "canal" else dp.seguimiento_metas(
            _df_mes_meta, _metas_all, mes_meta,
            dias_pasados=dias_pas, dias_totales=dias_tot,
            df_mes_anterior=_df_mes_prev, canales=_canales_vista,
            nivel=_nivel,
            desde=_desde_m, corte=_corte, hasta=_ult_dia_m,
        )

        _obj_canal = float(_seg_canal["meta_kg"].sum()) if not _seg_canal.empty else 0.0
        _obj_nivel = float(_seg["meta_kg"].sum()) if not _seg.empty else 0.0
        # Si todavía no se cargó el total del canal, se cae a la suma del nivel
        # que se está mirando para no dejar la solapa muda.
        _obj = _obj_canal if _obj_canal > 0 else _obj_nivel
        _av = float(_seg_canal["avance_kg"].sum()) if not _seg_canal.empty else 0.0
        _proy = float(_seg_canal["proyeccion_kg"].sum()) if not _seg_canal.empty else 0.0
        _prev_kg = float(_seg_canal["mes_ant_kg"].sum()) if not _seg_canal.empty else 0.0

        if _obj <= 0:
            st.info(
                f"Todavía no hay objetivos cargados para {etiqueta_mes(mes_meta)}"
                + (f" en {canal_meta}." if canal_meta != "TODOS" else ".")
                + " Cargalos en la solapa **Cargar objetivos**."
            )
        else:
            _alc = _proy / _obj * 100 if _obj else 0

            k1, k2, k3, k4 = st.columns(4)
            k1.metric(
                "Objetivo del mes", fmt_kg(_obj),
                help=("Total cargado a nivel canal."
                      if _obj_canal > 0 else
                      "Todavía no se cargó el total del canal: se está "
                      f"mostrando la suma del nivel {_nivel_lbl.lower()}."),
            )
            k2.metric(
                "Avance real", fmt_kg(_av),
                delta=f"{_av / _obj * 100:,.0f}% del objetivo".replace(",", ".")
                if _obj else None, delta_color="off",
            )
            k3.metric(
                "Proyección a fin de mes", fmt_kg(_proy),
                delta=fmt_kg(_proy - _obj) + " vs objetivo",
                delta_color="normal",
                help=("Cómo va a cerrar el mes si se mantiene el ritmo actual. "
                      "Se calcula vendedor por vendedor —avance ÷ días de "
                      "venta transcurridos × días del mes— y se suma: en Food "
                      "Service cada vendedor cuenta sus días de facturación "
                      "(factura dos a cuatro días fijos por semana), en el "
                      "resto de los canales son días hábiles de lunes a "
                      f"sábado. En promedio, día {dias_pas} de {dias_tot}."
                      if not _cerrado else
                      "El mes ya cerró: la proyección es el kilaje real."),
            )
            k4.metric(
                "Alcance proyectado", f"{_alc:,.1f} %".replace(",", "."),
                delta=(f"{(_av / _prev_kg - 1) * 100:+,.1f}% vs {etiqueta_mes(_mes_prev)}"
                       .replace(",", ".")) if _prev_kg else None,
                delta_color="normal",
            )

            _falta = max(_obj - _av, 0.0)
            _dias_rest = max(dias_tot - dias_pas, 0)
            if _falta > 0 and _dias_rest > 0:
                # En Food el "día" que queda es un día de facturación, que es
                # el que importa: no sirve saber cuántos kg/día hacen falta si
                # el vendedor solo factura lunes y jueves.
                _uni, _unis = (("día de facturación", "días de facturación")
                               if _dias_mixto else ("día", "días de venta"))
                st.caption(
                    f"Faltan **{fmt_kg(_falta)}** para llegar al objetivo: "
                    f"**{fmt_kg(_falta / _dias_rest)}/{_uni}** en los "
                    f"{_dias_rest} {_unis} que quedan."
                )
            elif _falta <= 0:
                st.caption("Objetivo del mes ya cubierto con las ventas reales.")
            else:
                st.caption(
                    f"El mes cerró **{fmt_kg(_obj - _av)}** por debajo del objetivo."
                )

            # --- Controles de consistencia --------------------------------
            # No bloquean nada: avisan. La idea es que haya una lógica detrás
            # de los números y no campos sueltos donde cargar valores.
            _vals = []
            for _c in _canales_vista:
                _v = dp.validar_metas(_metas_all, mes_meta, _c)
                if not _v.empty:
                    _v.insert(0, "Canal", _c)
                    _vals.append(_v)
            _val = pd.concat(_vals, ignore_index=True) if _vals else pd.DataFrame()

            if not _val.empty:
                st.divider()
                _n_av = int((_val["estado"] == "aviso").sum())
                _n_falta = int((_val["estado"] == "falta").sum())

                _vv = _val.copy()
                _vv.insert(0, "", _vv["estado"].map(dp.icono_validacion))
                _vv = _vv.drop(columns=["estado"]).rename(columns={
                    "control": "Control",
                    "esperado_kg": "Esperado (kg)",
                    "cargado_kg": "Cargado (kg)",
                    "dif_kg": "Diferencia (kg)",
                    "dif_pct": "Dif. %",
                    "detalle": "Detalle",
                })
                if canal_meta != "TODOS":
                    _vv = _vv.drop(columns=["Canal"])

                _cfg_val = {
                    "": st.column_config.TextColumn(width="small"),
                    "Esperado (kg)": st.column_config.NumberColumn(format="%.0f"),
                    "Cargado (kg)": st.column_config.NumberColumn(format="%.0f"),
                    "Diferencia (kg)": st.column_config.NumberColumn(format="%.0f"),
                    "Dif. %": st.column_config.NumberColumn(format="%.1f %%"),
                    "Detalle": st.column_config.TextColumn(width="large"),
                }

                if canal_meta == "TODOS":
                    _resumen = (
                        f"{_n_av} desvío(s) y {_n_falta} dato(s) faltante(s) "
                        "entre todos los canales")
                    with st.expander(f"Controles de consistencia · {_resumen}",
                                     expanded=_n_av > 0):
                        st.dataframe(_vv, use_container_width=True,
                                     hide_index=True, column_config=_cfg_val)
                else:
                    st.markdown("#### Controles de consistencia")
                    if _n_av:
                        st.warning(
                            f"{_n_av} control(es) no cierran. Los objetivos "
                            "están cargados pero no suman lo mismo entre "
                            "niveles.")
                    elif _n_falta:
                        st.info("Falta cargar niveles para poder validar la "
                                "apertura completa.")
                    else:
                        st.success("Los objetivos cierran entre los tres niveles.")
                    st.dataframe(_vv, use_container_width=True, hide_index=True,
                                 column_config=_cfg_val)

            st.divider()

            # --- Tabla de seguimiento ----------------------------------------
            _t = _seg.copy()
            _t.insert(0, "sem", _t["alcance_pct"].map(dp.semaforo))
            _vista = _t.rename(columns={
                "sem": "",
                "dsCanalMkt": "Canal",
                "marca_linea": "Marca / Línea",
                "dsVendedor": "Vendedor",
                "meta_kg": "Objetivo (kg)",
                "avance_kg": "Avance (kg)",
                "proyeccion_kg": "Proyección (kg)",
                "alcance_pct": "Alcance %",
                "brecha_kg": "Brecha (kg)",
                "falta_kg": "Falta (kg)",
                "mes_ant_kg": "Mes anterior (kg)",
                "var_ant_pct": "Var. vs mes ant. %",
            })
            if canal_meta != "TODOS" and _nivel != "canal":
                _vista = _vista.drop(columns=["Canal"])

            st.dataframe(
                _vista,
                use_container_width=True, hide_index=True,
                column_config={
                    "": st.column_config.TextColumn(width="small"),
                    "Objetivo (kg)": st.column_config.NumberColumn(format="%.0f"),
                    "Avance (kg)": st.column_config.NumberColumn(format="%.0f"),
                    "Proyección (kg)": st.column_config.NumberColumn(format="%.0f"),
                    "Alcance %": st.column_config.NumberColumn(format="%.1f %%"),
                    "Brecha (kg)": st.column_config.NumberColumn(format="%.0f"),
                    "Falta (kg)": st.column_config.NumberColumn(format="%.0f"),
                    "Mes anterior (kg)": st.column_config.NumberColumn(format="%.0f"),
                    "Var. vs mes ant. %": st.column_config.NumberColumn(format="%.1f %%"),
                },
            )
            st.caption(
                "🟢 proyecta llegar al objetivo · 🟡 entre 90% y 100% · "
                "🔴 por debajo del 90% · ⚪ vendió sin meta cargada."
            )

            # Resumen por canal (solo tiene sentido en la vista consolidada)
            if (canal_meta == "TODOS" and _nivel != "canal"
                    and _seg["dsCanalMkt"].nunique() > 1):
                with st.expander("Resumen por canal"):
                    _porc = (
                        _seg.groupby("dsCanalMkt", as_index=False)
                        .agg(meta_kg=("meta_kg", "sum"),
                             avance_kg=("avance_kg", "sum"),
                             proyeccion_kg=("proyeccion_kg", "sum"),
                             mes_ant_kg=("mes_ant_kg", "sum"))
                    )
                    _porc["alcance_pct"] = np.where(
                        _porc["meta_kg"] > 0,
                        _porc["proyeccion_kg"] / _porc["meta_kg"] * 100, np.nan)
                    _porc.insert(0, "sem", _porc["alcance_pct"].map(dp.semaforo))
                    st.dataframe(
                        _porc.rename(columns={
                            "sem": "", "dsCanalMkt": "Canal",
                            "meta_kg": "Objetivo (kg)", "avance_kg": "Avance (kg)",
                            "proyeccion_kg": "Proyección (kg)",
                            "alcance_pct": "Alcance %",
                            "mes_ant_kg": "Mes anterior (kg)"}),
                        use_container_width=True, hide_index=True,
                    )

            _hojas = {"Seguimiento de metas": _vista}
            if not _val.empty:
                _hojas["Controles"] = _vv
            boton_excel("metas", _hojas, key="xlsx_metas")

    # =====================================================================
    # EVOLUTIVO (proyectado vs. meta a lo largo del año)
    # =====================================================================
    # El Seguimiento de arriba mira UN mes. Acá se ve la curva del año:
    # objetivo vs. vendido mes a mes, con el mes abierto proyectado a fin de
    # mes. Es la lectura que se pidió para la mesa chica ("el proyectado
    # contra la meta, cómo viene de cumplimiento").
    #
    # No depende del selector de mes de arriba (muestra el año entero) pero sí
    # del de canal, igual que el resto de la solapa. Solo se compara contra el
    # OBJETIVO cargado: los meses sin objetivo muestran la venta real y no
    # calculan cumplimiento (no se los completa con el presupuesto anual).
    with sub_evol:
        # etiqueta -> (nivel de la meta, columna del ítem, nombre corto). El
        # nombre corto es para la hoja del Excel: openpyxl no acepta "/" en el
        # título de una hoja y "Proveedor / línea" la rompía.
        _NIVEL_EVOL = {
            "Total empresa": ("canal", None, "total"),
            "Canal": ("canal", "dsCanalMkt", "canal"),
            "Proveedor / línea": ("proveedor", "marca_linea", "proveedor"),
            "Vendedor": ("vendedor", "dsVendedor", "vendedor"),
        }
        _lbl_evol = st.radio(
            "Abrir por", list(_NIVEL_EVOL.keys()), horizontal=True,
            key="metas_nivel_evol",
            help="El gráfico de barras es siempre el total de lo que estés "
                 "mirando. La apertura cambia el detalle de abajo.",
        )
        _niv_evol, _col_evol, _corto_evol = _NIVEL_EVOL[_lbl_evol]

        _ev = cargar_evolutivo(
            _niv_evol, tuple(_canales_vista),
            os.path.getmtime(PARQUET_PATH), _mtime_acuerdos(), _mtime_metas(),
            _hoy,
        )
        _ev_tot = dp.evolutivo_total(_ev)

        if _ev_tot.empty:
            st.info(
                "Todavía no hay meses con ventas para armar el evolutivo."
            )
        else:
            _mes_lbl = {m: MESES_ES[int(m[5:7]) - 1][:3] for m in _ev_tot["anio_mes"]}
            _orden_mes = list(_ev_tot["anio_mes"])
            _abierto = _ev_tot[~_ev_tot["cerrado"]]["anio_mes"].tolist()

            # --- Acumulado del año -------------------------------------------
            # Solo entran los meses CON objetivo cargado: sumar el real de un
            # mes sin meta contra un objetivo que no existe daría un
            # cumplimiento inflado.
            _con_meta = _ev_tot[_ev_tot["meta_kg"] > 0]
            _meta_ac = float(_con_meta["meta_kg"].sum())
            _proy_ac = float(_con_meta["proyeccion_kg"].sum())
            _real_ac = float(_con_meta["real_kg"].sum())

            if _con_meta.empty:
                st.warning(
                    "No hay ningún objetivo cargado en "
                    f"{dp.ANIO}: el evolutivo muestra la venta real, pero sin "
                    "línea de meta ni cumplimiento. Cargalos en "
                    "**Cargar objetivos**."
                )
            else:
                e1, e2, e3, e4 = st.columns(4)
                e1.metric(
                    "Meses con objetivo", f"{len(_con_meta)}",
                    help="Meses del año que tienen objetivo cargado. El "
                         "acumulado de al lado se calcula solo sobre estos.",
                )
                e2.metric("Objetivo acumulado", fmt_kg(_meta_ac))
                e3.metric(
                    "Vendido + proyectado", fmt_kg(_proy_ac),
                    delta=fmt_kg(_proy_ac - _meta_ac) + " vs objetivo",
                    help=f"Real acumulado: {fmt_kg(_real_ac)}. El mes abierto "
                         "entra proyectado a fin de mes.",
                )
                e4.metric(
                    "Cumplimiento acumulado",
                    f"{_proy_ac / _meta_ac * 100:,.1f} %".replace(",", ".")
                    if _meta_ac else "—",
                )

            # --- Gráfico: barras vendido + proyectado, línea de objetivo -----
            _b = _ev_tot.copy()
            _b["Vendido"] = _b["real_kg"]
            _b["Proyectado"] = (_b["proyeccion_kg"] - _b["real_kg"]).clip(lower=0)
            _b["mes"] = _b["anio_mes"].map(_mes_lbl)
            _long = _b.melt(
                id_vars=["anio_mes", "mes"], value_vars=["Vendido", "Proyectado"],
                var_name="Tramo", value_name="kg",
            )

            fig_ev = px.bar(
                _long, x="mes", y="kg", color="Tramo",
                category_orders={"mes": [_mes_lbl[m] for m in _orden_mes],
                                 "Tramo": ["Vendido", "Proyectado"]},
                color_discrete_map={"Vendido": "#2a8ed4",
                                    "Proyectado": "#3f5b7a"},
            )
            # Los meses sin objetivo quedan como hueco en la línea
            # (connectgaps=False): no se inventa una meta que no se cargó.
            _meta_y = _ev_tot["meta_kg"].where(_ev_tot["meta_kg"] > 0)
            if _meta_y.notna().any():
                fig_ev.add_scatter(
                    x=[_mes_lbl[m] for m in _orden_mes], y=_meta_y,
                    mode="lines+markers", name="Objetivo", connectgaps=False,
                    line=dict(color="#f59e0b", width=3, dash="dash"),
                    marker=dict(size=9),
                )
            fig_ev.update_layout(
                barmode="stack",
                template="plotly_dark",
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                margin=dict(l=10, r=10, t=10, b=10),
                legend=dict(title="", orientation="h", y=-0.2),
                xaxis_title="Mes",
                yaxis_title="Kilos",
                height=420,
            )
            st.plotly_chart(fig_ev, use_container_width=True)
            st.caption(
                ("Mes abierto: **" + etiqueta_mes(_abierto[0]) + "** — la "
                 "parte clara de la barra es lo que falta para cerrar el mes "
                 "al ritmo actual (misma proyección que el Seguimiento). "
                 if _abierto else "Todos los meses están cerrados. ")
                + "Los meses sin objetivo cargado no cortan la línea naranja: "
                "muestran la venta real y no calculan cumplimiento."
            )

            # NOTA: acá iba un gráfico de líneas con el % de cumplimiento mes
            # a mes (una línea por canal / proveedor / vendedor). Se sacó a
            # pedido: el cumplimiento ya se lee en la tabla de abajo y en la
            # grilla ítem × mes, y con varias líneas encimadas el gráfico
            # aportaba poco. El dato sigue estando en dp.evolutivo_metas().

            # --- Tablas ------------------------------------------------------
            st.divider()
            _t_ev = _ev_tot.copy()
            _t_ev.insert(0, "sem", _t_ev["cumplimiento_pct"].map(dp.semaforo))
            _t_ev["Mes"] = [
                etiqueta_mes(m) + ("" if cer else " · abierto")
                for m, cer in zip(_t_ev["anio_mes"], _t_ev["cerrado"])
            ]
            _t_ev = _t_ev[["sem", "Mes", "meta_kg", "real_kg", "proyeccion_kg",
                           "cumplimiento_pct", "brecha_kg"]].rename(columns={
                "sem": "",
                "meta_kg": "Objetivo (kg)",
                "real_kg": "Vendido (kg)",
                "proyeccion_kg": "Proyección (kg)",
                "cumplimiento_pct": "Cumplimiento %",
                "brecha_kg": "Desvío (kg)",
            })
            st.dataframe(
                _t_ev, use_container_width=True, hide_index=True,
                column_config={
                    "": st.column_config.TextColumn(width="small"),
                    "Objetivo (kg)": st.column_config.NumberColumn(format="%.0f"),
                    "Vendido (kg)": st.column_config.NumberColumn(format="%.0f"),
                    "Proyección (kg)": st.column_config.NumberColumn(format="%.0f"),
                    "Cumplimiento %": st.column_config.NumberColumn(format="%.1f %%"),
                    "Desvío (kg)": st.column_config.NumberColumn(format="%.0f"),
                },
            )
            st.caption(
                "🟢 proyecta llegar al objetivo · 🟡 entre 90% y 100% · "
                "🔴 por debajo del 90% · ⚪ sin objetivo cargado ese mes."
            )

            _hojas_ev = {"Evolutivo (total)": _t_ev}

            # Detalle ítem × mes: la grilla que se mira en la reunión.
            _piv = None
            if _col_evol is not None:
                _d = _ev[_ev["meta_kg"] > 0]
                if not _d.empty:
                    _d = (_d.groupby(["anio_mes", _col_evol], as_index=False)
                          .agg(meta_kg=("meta_kg", "sum"),
                               proyeccion_kg=("proyeccion_kg", "sum")))
                    _d["cumplimiento_pct"] = (_d["proyeccion_kg"]
                                              / _d["meta_kg"] * 100)
                    _piv = (_d.pivot(index=_col_evol, columns="anio_mes",
                                     values="cumplimiento_pct")
                            .rename(columns=_mes_lbl))
                    _piv = _piv[[c for c in
                                 [_mes_lbl[m] for m in _orden_mes]
                                 if c in _piv.columns]]
                    _piv.index.name = _lbl_evol
                    with st.expander(
                            f"Cumplimiento % por {_lbl_evol.lower()} y mes",
                            expanded=True):
                        st.dataframe(
                            _piv.style.format("{:.0f} %", na_rep="—"),
                            use_container_width=True,
                        )
                        st.caption(
                            "Vacío = ese mes no tenía objetivo cargado para "
                            f"ese {_lbl_evol.lower()}."
                        )
                    _hojas_ev[f"Cumplimiento por {_corto_evol}"] = (
                        _piv.reset_index())

            boton_excel("evolutivo_metas", _hojas_ev, key="xlsx_evolutivo")

    # =====================================================================
    # CARGA DE OBJETIVOS (canal → proveedor → vendedor)
    # =====================================================================
    with sub_carga:
        if canal_meta == "TODOS":
            st.info(
                "Elegí un canal arriba para cargar los objetivos. Se editan de "
                "a un canal por vez para no pisar lo que cargó otro supervisor."
            )
        else:
            st.caption(
                f"Objetivos de **{canal_meta}** para **{etiqueta_mes(mes_meta)}**, "
                "en kilos. El orden es: primero el total del canal, después la "
                "apertura por proveedor y el reparto entre vendedores. Cada "
                "bloque se guarda por separado y solo pisa su propio nivel."
            )

            def _ov_caption(nivel, etiqueta):
                """Muestra el objetivo original vs. el vigente cuando la meta se
                reajustó durante el mes."""
                _ov = dp.metas_original_vs_vigente(
                    mes_meta, "objetivo", nivel, canal_meta, historial=_hist_all)
                if _ov and _ov["n_cargas"] > 1 and abs(
                        _ov["vigente_kg"] - _ov["original_kg"]) > 1:
                    st.caption(
                        f"{etiqueta}: original {fmt_kg(_ov['original_kg'])} "
                        f"({_ov['fecha_original']:%d/%m %H:%M}) → vigente "
                        f"{fmt_kg(_ov['vigente_kg'])} "
                        f"({_ov['fecha_vigente']:%d/%m %H:%M}, "
                        f"{_ov['n_cargas']} cargas)."
                    )

            # ---------------------------------------------------------------
            # 1. Objetivo total del canal
            # ---------------------------------------------------------------
            st.markdown("#### 1. Objetivo total del canal")
            _obj_canal_act = dp.total_meta(
                _metas_all, mes_meta, "objetivo", "canal", canal_meta)
            _pres_mes = dp.total_meta(
                _metas_all, mes_meta, "presupuesto", "canal", canal_meta)
            _real_prev_canal = float(
                dp.kilos_por_canal(_df_mes_prev)
                .query("dsCanalMkt == @canal_meta")["kilos"].sum())

            cc1, cc2, cc3 = st.columns([1.6, 1, 1])
            _nuevo_canal = cc1.number_input(
                f"Kilos a vender en {canal_meta}",
                min_value=0.0, step=1000.0, format="%.0f",
                value=float(_obj_canal_act),
                key=f"meta_canal_{mes_meta}_{canal_meta}",
                help="Es el número por el que responde el supervisor del canal. "
                     "La apertura por proveedor y por vendedor se valida contra "
                     "este total.",
            )
            cc2.metric(
                etiqueta_mes(_mes_prev), fmt_kg(_real_prev_canal),
                help="Kilos que el canal vendió realmente el mes pasado. "
                     "Referencia para dimensionar el objetivo.",
            )
            cc3.metric(
                "Presupuesto del mes",
                fmt_kg(_pres_mes) if _pres_mes > 0 else "—",
                delta=(f"{(_nuevo_canal / _pres_mes - 1) * 100:+,.1f}%".replace(",", ".")
                       if _pres_mes > 0 and _nuevo_canal > 0 else None),
                delta_color="off",
                help="Lo que se estimó a principio de año para este mes. El "
                     "objetivo mensual puede desviarse a propósito.",
            )
            _ov_caption("canal", "Objetivo del canal")

            if st.button("Guardar objetivo del canal", type="primary",
                         key="btn_meta_canal"):
                _fila = pd.DataFrame([{
                    "anio_mes": mes_meta, "dsCanalMkt": canal_meta,
                    "marca_linea": "", "dsVendedor": "",
                    "meta_kg": float(_nuevo_canal),
                }])
                dp.upsert_metas(_fila, anio_mes=mes_meta, canales=[canal_meta],
                                tipo="objetivo", nivel="canal")
                st.session_state["_metas_ok"] = (
                    f"Objetivo de {canal_meta} para {etiqueta_mes(mes_meta)}: "
                    f"{fmt_kg(_nuevo_canal)}."
                )
                st.rerun()

            st.divider()

            # ---------------------------------------------------------------
            # 2. Apertura por proveedor / línea
            # ---------------------------------------------------------------
            st.markdown("#### 2. Apertura por proveedor / línea")

            # Filas base: marcas con meta cargada + las que vendieron en el mes o
            # en el anterior. Así la grilla arranca con el universo real de marcas
            # del canal y no hay que tipear nombres.
            _marcas_canal = sorted(set(
                dp.kilos_por_canal_marca(_df_mes_meta)
                .query("dsCanalMkt == @canal_meta")["marca_linea"]
            ) | set(
                dp.kilos_por_canal_marca(_df_mes_prev)
                .query("dsCanalMkt == @canal_meta")["marca_linea"]
            ))
            _marcas_todas = sorted(set(_marcas_canal) | set(
                _df_full_metas["marca_linea"].dropna().astype(str).str.strip()
                .replace({"": None}).dropna().unique().tolist()
            ))

            _cargadas = dp.filtrar_metas(
                _metas_all, anio_mes=mes_meta, tipo="objetivo",
                nivel="proveedor", canal=canal_meta)[["marca_linea", "meta_kg"]]

            _ref = (
                dp.kilos_por_canal_marca(_df_mes_prev)
                .query("dsCanalMkt == @canal_meta")[["marca_linea", "kilos"]]
                .rename(columns={"kilos": "_ref"})
            )
            _base = pd.DataFrame({"marca_linea": _marcas_canal})
            _base = _base.merge(_cargadas, on="marca_linea", how="outer")
            _base = _base.merge(_ref, on="marca_linea", how="left")
            _base["meta_kg"] = pd.to_numeric(
                _base["meta_kg"], errors="coerce").fillna(0.0)
            _base["_ref"] = pd.to_numeric(_base["_ref"], errors="coerce").fillna(0.0)
            _base = _base.sort_values(
                ["meta_kg", "_ref"], ascending=False).reset_index(drop=True)

            _col_ref = f"{etiqueta_mes(_mes_prev)} (kg)"
            _grilla = _base.rename(columns={
                "marca_linea": "Marca / Línea",
                "meta_kg": "Meta (kg)",
                "_ref": _col_ref,
            })

            _edit = st.data_editor(
                _grilla,
                key=f"metas_editor_prov_{mes_meta}_{canal_meta}",
                use_container_width=True, hide_index=True, num_rows="dynamic",
                column_config={
                    "Marca / Línea": st.column_config.SelectboxColumn(
                        options=_marcas_todas, required=True, width="large"),
                    "Meta (kg)": st.column_config.NumberColumn(
                        min_value=0.0, step=100.0, format="%.0f",
                        help="Kilos que se quieren vender en el mes. "
                             "Dejar en 0 borra la meta de esa marca."),
                    _col_ref: st.column_config.NumberColumn(
                        format="%.0f", disabled=True,
                        help="Kilos reales del mes anterior. Referencia, no se "
                             "guarda."),
                },
            )

            _tot_prov = pd.to_numeric(
                _edit["Meta (kg)"], errors="coerce").fillna(0).sum()
            _tot_ref = pd.to_numeric(
                _edit[_col_ref], errors="coerce").fillna(0).sum()

            _p1, _p2, _p3 = st.columns([1.2, 1.2, 2])
            _p1.metric(
                etiqueta_mes(mes_meta), fmt_kg(_tot_prov),
                help="Suma de la meta que estás cargando en la grilla de "
                     "arriba, para el mes de la meta.",
            )
            _p2.metric(
                etiqueta_mes(_mes_prev), fmt_kg(_tot_ref),
                delta=(f"{(_tot_prov / _tot_ref - 1) * 100:+,.1f}%".replace(",", ".")
                       if _tot_ref else None),
                delta_color="off",
                help="Kilos que estas marcas vendieron realmente el mes "
                     "pasado. El % es cuánto más (o menos) estás pidiendo.",
            )
            # Control en vivo contra el total del canal, antes de guardar.
            if _nuevo_canal > 0:
                _dif_p = _tot_prov - _nuevo_canal
                _p3.metric(
                    "Diferencia vs objetivo del canal", fmt_kg(_dif_p),
                    delta="cierra" if abs(_dif_p) <= max(_nuevo_canal * 0.005, 1)
                    else "no cierra",
                    delta_color="off",
                )

            if st.button("Guardar apertura por proveedor", type="primary",
                         key="btn_guardar_metas_prov"):
                _nuevas = _edit.rename(columns={
                    "Marca / Línea": "marca_linea", "Meta (kg)": "meta_kg"})
                _nuevas = _nuevas[["marca_linea", "meta_kg"]].copy()
                _nuevas["anio_mes"] = mes_meta
                _nuevas["dsCanalMkt"] = canal_meta
                _nuevas["dsVendedor"] = ""
                _guardadas = dp.upsert_metas(
                    _nuevas, anio_mes=mes_meta, canales=[canal_meta],
                    tipo="objetivo", nivel="proveedor")
                _n = len(dp.filtrar_metas(
                    _guardadas, anio_mes=mes_meta, tipo="objetivo",
                    nivel="proveedor", canal=canal_meta))
                st.session_state["_metas_ok"] = (
                    f"Guardadas {_n} metas por proveedor de {canal_meta} para "
                    f"{etiqueta_mes(mes_meta)} ({fmt_kg(_tot_prov)} en total)."
                )
                st.rerun()

            _ov_caption("proveedor", "Apertura por proveedor")

            st.divider()

            # ---------------------------------------------------------------
            # 3. Reparto entre vendedores
            # ---------------------------------------------------------------
            st.markdown("#### 3. Reparto entre vendedores")
            st.caption(
                "El objetivo del vendedor es sobre el total del canal, no por "
                "proveedor: abrir los tres ejes a la vez da una grilla "
                "inmanejable y el control de que la suma cierre se cumple igual."
            )

            _vend_canal = sorted(set(
                dp.kilos_por_canal_vendedor(_df_mes_meta)
                .query("dsCanalMkt == @canal_meta")["dsVendedor"]
            ) | set(
                dp.kilos_por_canal_vendedor(_df_mes_prev)
                .query("dsCanalMkt == @canal_meta")["dsVendedor"]
            ))
            _vend_todos = sorted(set(_vend_canal) | set(
                _df_full_metas["dsVendedor"].dropna().astype(str).str.strip()
                .replace({"": None}).dropna().unique().tolist()
            ))

            _cargados_v = dp.filtrar_metas(
                _metas_all, anio_mes=mes_meta, tipo="objetivo",
                nivel="vendedor", canal=canal_meta)[["dsVendedor", "meta_kg"]]

            _ref_v = (
                dp.kilos_por_canal_vendedor(_df_mes_prev)
                .query("dsCanalMkt == @canal_meta")[["dsVendedor", "kilos"]]
                .rename(columns={"kilos": "_ref"})
            )
            _base_v = pd.DataFrame({"dsVendedor": _vend_canal})
            _base_v = _base_v.merge(_cargados_v, on="dsVendedor", how="outer")
            _base_v = _base_v.merge(_ref_v, on="dsVendedor", how="left")
            _base_v["meta_kg"] = pd.to_numeric(
                _base_v["meta_kg"], errors="coerce").fillna(0.0)
            _base_v["_ref"] = pd.to_numeric(
                _base_v["_ref"], errors="coerce").fillna(0.0)
            _base_v = _base_v.sort_values(
                ["meta_kg", "_ref"], ascending=False).reset_index(drop=True)

            _grilla_v = _base_v.rename(columns={
                "dsVendedor": "Vendedor",
                "meta_kg": "Meta (kg)",
                "_ref": _col_ref,
            })

            _edit_v = st.data_editor(
                _grilla_v,
                key=f"metas_editor_vend_{mes_meta}_{canal_meta}",
                use_container_width=True, hide_index=True, num_rows="dynamic",
                column_config={
                    "Vendedor": st.column_config.SelectboxColumn(
                        options=_vend_todos, required=True, width="large"),
                    "Meta (kg)": st.column_config.NumberColumn(
                        min_value=0.0, step=100.0, format="%.0f",
                        help="Kilos asignados al vendedor en el mes. "
                             "Dejar en 0 borra su meta."),
                    _col_ref: st.column_config.NumberColumn(
                        format="%.0f", disabled=True,
                        help="Kilos reales del mes anterior. Referencia, no se "
                             "guarda."),
                },
            )

            _tot_vend = pd.to_numeric(
                _edit_v["Meta (kg)"], errors="coerce").fillna(0).sum()
            _tot_ref_v = pd.to_numeric(
                _edit_v[_col_ref], errors="coerce").fillna(0).sum()

            _v1, _v2, _v3 = st.columns([1.2, 1.2, 2])
            _v1.metric(
                etiqueta_mes(mes_meta), fmt_kg(_tot_vend),
                help="Suma de lo que estás repartiendo entre los vendedores, "
                     "para el mes de la meta.",
            )
            _v2.metric(
                etiqueta_mes(_mes_prev), fmt_kg(_tot_ref_v),
                delta=(f"{(_tot_vend / _tot_ref_v - 1) * 100:+,.1f}%".replace(",", ".")
                       if _tot_ref_v else None),
                delta_color="off",
                help="Kilos que estos vendedores facturaron realmente el mes "
                     "pasado. El % es cuánto más (o menos) estás pidiendo.",
            )
            if _nuevo_canal > 0:
                _dif_v = _tot_vend - _nuevo_canal
                _v3.metric(
                    "Diferencia vs objetivo del canal", fmt_kg(_dif_v),
                    delta="cierra" if abs(_dif_v) <= max(_nuevo_canal * 0.005, 1)
                    else "no cierra",
                    delta_color="off",
                )

            if st.button("Guardar reparto entre vendedores", type="primary",
                         key="btn_guardar_metas_vend"):
                _nuevas_v = _edit_v.rename(columns={
                    "Vendedor": "dsVendedor", "Meta (kg)": "meta_kg"})
                _nuevas_v = _nuevas_v[["dsVendedor", "meta_kg"]].copy()
                _nuevas_v["anio_mes"] = mes_meta
                _nuevas_v["dsCanalMkt"] = canal_meta
                _nuevas_v["marca_linea"] = ""
                _guardadas_v = dp.upsert_metas(
                    _nuevas_v, anio_mes=mes_meta, canales=[canal_meta],
                    tipo="objetivo", nivel="vendedor")
                _nv = len(dp.filtrar_metas(
                    _guardadas_v, anio_mes=mes_meta, tipo="objetivo",
                    nivel="vendedor", canal=canal_meta))
                st.session_state["_metas_ok"] = (
                    f"Guardadas {_nv} metas de vendedores de {canal_meta} para "
                    f"{etiqueta_mes(mes_meta)} ({fmt_kg(_tot_vend)} en total)."
                )
                st.rerun()

            _ov_caption("vendedor", "Reparto entre vendedores")

    # =====================================================================
    # PRESUPUESTO ANUAL
    # =====================================================================
    # Es la línea base que se define a principio de año y NO se toca. El
    # objetivo mensual se compara contra esto para ver el desvío acumulado.
    with sub_pres:
        if canal_meta == "TODOS":
            st.info(
                "Elegí un canal arriba para cargar su presupuesto anual."
            )
        else:
            st.caption(
                "Lo que la empresa estima vender en el año, mes por mes. Es la "
                "línea base: se carga una vez y queda fija. El objetivo mensual "
                "se carga aparte y puede desviarse de acá a propósito."
            )

            _anios_datos = sorted({m[:4] for m in _meses_disp}, reverse=True)
            _anio_mes_meta = mes_meta[:4]
            _anios_opts = sorted(
                set(_anios_datos) | {_anio_mes_meta,
                                     str(int(_anio_mes_meta) + 1)},
                reverse=True)
            anio_pres = st.selectbox(
                "Año", _anios_opts,
                index=_anios_opts.index(_anio_mes_meta),
                key="pres_anio",
            )

            _meses_anio = [f"{anio_pres}-{m:02d}" for m in range(1, 13)]

            _pres_cargado = dp.filtrar_metas(
                _metas_all, tipo="presupuesto", nivel="canal", canal=canal_meta)
            _pres_cargado = _pres_cargado[
                _pres_cargado["anio_mes"].isin(_meses_anio)
            ][["anio_mes", "meta_kg"]]

            _real_mes = dp.kilos_por_mes_canal(_df_full_metas)
            _real_mes = _real_mes[_real_mes["dsCanalMkt"] == canal_meta]
            _real_map = dict(zip(_real_mes["anio_mes"], _real_mes["kilos"]))

            _bp = pd.DataFrame({"anio_mes": _meses_anio})
            _bp = _bp.merge(_pres_cargado, on="anio_mes", how="left")
            _bp["meta_kg"] = pd.to_numeric(
                _bp["meta_kg"], errors="coerce").fillna(0.0)
            # Sin columna del año anterior: el parquet arranca en el año en
            # curso, así que esa referencia daba siempre 0 y solo ocupaba lugar.
            _bp["_real"] = _bp["anio_mes"].map(
                lambda am: float(_real_map.get(am, 0.0)))
            _bp["Mes"] = _bp["anio_mes"].map(etiqueta_mes)

            _col_real = f"Real {anio_pres} (kg)"
            _grilla_p = _bp[["Mes", "meta_kg", "_real"]].rename(
                columns={"meta_kg": "Presupuesto (kg)", "_real": _col_real})

            _edit_p = st.data_editor(
                _grilla_p,
                key=f"pres_editor_{anio_pres}_{canal_meta}",
                use_container_width=True, hide_index=True, num_rows="fixed",
                column_config={
                    "Mes": st.column_config.TextColumn(disabled=True),
                    "Presupuesto (kg)": st.column_config.NumberColumn(
                        min_value=0.0, step=1000.0, format="%.0f",
                        help="Kilos estimados para ese mes. Dejar en 0 lo borra."),
                    _col_real: st.column_config.NumberColumn(
                        format="%.0f", disabled=True,
                        help="Kilos reales de este año (los meses ya cerrados). "
                             "Referencia, no se guarda."),
                },
            )

            _tot_pres = pd.to_numeric(
                _edit_p["Presupuesto (kg)"], errors="coerce").fillna(0).sum()
            _tot_real = pd.to_numeric(
                _edit_p[_col_real], errors="coerce").fillna(0).sum()

            _q1, _q2 = st.columns(2)
            _q1.metric(f"Presupuesto {anio_pres}", fmt_kg(_tot_pres))
            _q2.metric(
                f"Real acumulado {anio_pres}", fmt_kg(_tot_real),
                delta=(f"{_tot_real / _tot_pres * 100:,.1f}% del presupuesto"
                       .replace(",", ".") if _tot_pres else None),
                delta_color="off",
                help="Kilos ya facturados en el año, contra el presupuesto "
                     "de los 12 meses.",
            )

            if st.button("Guardar presupuesto anual", type="primary",
                         key="btn_guardar_pres"):
                # El mes se recupera desde la etiqueta, no por posición: si
                # Streamlit devolviera las filas reordenadas, el presupuesto
                # terminaría en el mes equivocado.
                _mapa_mes = dict(zip(_bp["Mes"], _bp["anio_mes"]))
                _np_ = _edit_p.rename(
                    columns={"Presupuesto (kg)": "meta_kg"})[
                        ["Mes", "meta_kg"]].copy()
                _np_["anio_mes"] = _np_["Mes"].map(_mapa_mes)
                _np_ = _np_.drop(columns=["Mes"])
                _np_["dsCanalMkt"] = canal_meta
                _np_["marca_linea"] = ""
                _np_["dsVendedor"] = ""
                dp.upsert_metas(
                    _np_, anio_mes=_meses_anio, canales=[canal_meta],
                    tipo="presupuesto", nivel="canal")
                st.session_state["_metas_ok"] = (
                    f"Presupuesto {anio_pres} de {canal_meta} guardado: "
                    f"{fmt_kg(_tot_pres)} en el año."
                )
                st.rerun()


# --- TAB ACUERDOS MCCAIN --------------------------------------------------
# Carga mensual del Excel de descuentos en el costo (cliente-artículo-mes).
# UPSERT POR MES: los meses que trae el Excel reemplazan a los guardados,
# los demás quedan intactos. El ajuste impacta en CM y CM% de todo el
# tablero al instante (el costo de Chess se guarda crudo y se ajusta al leer).
if tab_acuerdos is not None:
    with tab_acuerdos:
        st.subheader("Acuerdos comerciales McCain")
        st.caption(
            "Chess no incluye estos descuentos en el costo, así que el CM% "
            "sale más bajo de lo real. Subí acá el Excel de acuerdos de cada "
            "mes"
        )

        _msg_ok = st.session_state.pop("_acuerdos_ok", None)
        if _msg_ok:
            st.success(_msg_ok)

        # --- Estado actual (auditoría) ---------------------------------
        # OJO: acá se usa el dataset COMPLETO (df_full), no `df`, porque a
        # esta altura `df` ya viene filtrado por el período/filtros globales
        # y la auditoría debe mostrar el año entero sin importar qué mes
        # esté seleccionado arriba.
        _ac = dp.cargar_acuerdos()
        if _ac.empty:
            st.info("Todavía no hay acuerdos cargados.")
        else:
            _df_full = df_anio   # ver nota en "Altas y bajas": reusar, no recargar
            _aud = _df_full[_df_full["ajuste_mccain"] != 0].copy()
            _aud["Mes"] = _aud["fechaComprobate"].dt.to_period("M").astype(str)
            _por_mes = (
                _aud.groupby("Mes")
                .agg(**{
                    "Líneas ajustadas": ("ajuste_mccain", "size"),
                    "Ajuste aplicado $": ("ajuste_mccain", "sum"),
                })
                .reset_index()
            )
            _n_ac = (
                _ac.assign(Mes=lambda x: x["anio"].astype(str) + "-"
                           + x["mes"].astype(str).str.zfill(2))
                .groupby("Mes").size().rename("Acuerdos cargados")
                .reset_index()
            )
            _tabla = _n_ac.merge(_por_mes, on="Mes", how="left").fillna(0)
            _tabla["Líneas ajustadas"] = _tabla["Líneas ajustadas"].astype(int)
            _tabla["Ajuste aplicado $"] = _tabla["Ajuste aplicado $"].map(fmt_money)
            c1, c2 = st.columns([2, 1])
            c1.dataframe(_tabla, use_container_width=True, hide_index=True)
            c2.metric(
                "Ajuste total aplicado (todo el año)",
                fmt_money(_df_full["ajuste_mccain"].sum()),
                help="Importe restado del costo de Chess por acuerdos "
                     "McCain. Es la mejora directa de la contribución.",
            )

        st.divider()

        # --- Carga de un Excel nuevo ------------------------------------
        _subida = st.file_uploader(
            "Excel de acuerdos (.xlsx)", type=["xlsx", "xlsm"],
            key="up_acuerdos",
            help="Sirve tanto el acumulado del año como un archivo con "
                 "solo el mes nuevo.",
        )
        if _subida is not None:
            try:
                _nuevos, _res = dp.procesar_excel_acuerdos(_subida)
            except Exception as e:
                st.error(f"No pude procesar el archivo: {e}")
            else:
                st.markdown(
                    f"**{_res['filas_validas']}** acuerdos válidos "
                    f"(de {_res['filas_leidas']} filas leídas · "
                    f"{_res['descartadas']} vacías/incompletas descartadas · "
                    f"{_res['duplicados_resueltos']} duplicados resueltos "
                    "quedándose con el último)."
                )
                st.markdown(
                    "Meses que trae el archivo (van a **reemplazar** a los "
                    "cargados): " + ", ".join(_res["meses"])
                )
                if _res["negativos"]:
                    st.warning(
                        f"{_res['negativos']} acuerdos con valor NEGATIVO "
                        "(suben el costo en vez de bajarlo). Se aplican con "
                        "su signo; revisalos si no es intencional."
                    )
                    _neg = _nuevos[_nuevos["desc_kg"] < 0]
                    st.dataframe(_neg, use_container_width=True,
                                 hide_index=True)

                if st.button("Confirmar y aplicar al tablero",
                             type="primary", key="btn_acuerdos"):
                    dp.upsert_acuerdos(_nuevos)
                    st.cache_data.clear()  # recarga datos y serie ajustados
                    st.session_state["_acuerdos_ok"] = (
                        f"Listo: {_res['filas_validas']} acuerdos aplicados "
                        f"({', '.join(_res['meses'])}). CM y CM% ya están "
                        "recalculados en todo el tablero."
                    )
                    st.rerun()
